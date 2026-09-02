"""
Recálculo del libro con LibreOffice.

Para qué hace falta
-------------------
Cuando se escribe un valor en una celda, los resultados que quedaron
guardados en las celdas de fórmula son de antes del cambio. Excel los
vuelve a calcular al abrir el archivo, así que para el contador todo se ve
bien. El problema es nuestro: el programa necesita leer los totales para
mostrarlos en pantalla, y si los lee del archivo tal como quedó, lee
números viejos.

La solución es pasarle el archivo a LibreOffice, que sí puede recalcular
sin que nadie abra nada, y que vuelva a guardarlo con los resultados
adentro.

Si LibreOffice no está instalado, no pasa nada grave: el archivo se
entrega igual y el programa avisa que los totales solo se verán al abrirlo
en Excel. Nunca se rompe el flujo por esto.

Un detalle que costó encontrar
------------------------------
LibreOffice, de fábrica, NO recalcula al abrir un archivo de Excel: trae
puesto "nunca recalcular". Hay que decírselo con un ajuste de
configuración (OOXMLRecalcMode). Sin ese ajuste la conversión funciona,
no da ningún error, y devuelve el archivo con los totales viejos — que es
la peor forma de fallar, porque parece que funcionó.
"""

import shutil
import subprocess
import tempfile
import threading
import time
from pathlib import Path

from openpyxl import load_workbook

# Cuánto se espera a LibreOffice antes de darlo por perdido. Este libro
# tiene 902 fórmulas y en las pruebas tardó menos de 3 segundos, así que
# 120 sobra de lejos. Se deja alto igual: un computador lento con un libro
# más grande no debería fallar por apuro.
TIEMPO_LIMITE = 120

# Dónde suele quedar instalado LibreOffice en cada sistema. Se prueban en
# orden y se usa el primero que exista.
RUTAS_CONOCIDAS = (
    # Mac
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    # Windows
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    # Linux
    "/usr/bin/soffice",
    "/usr/bin/libreoffice",
    "/snap/bin/libreoffice",
)

# Los errores que Excel muestra dentro de una celda. Si aparece uno de
# estos después de recalcular, algo se rompió.
ERRORES_DE_EXCEL = frozenset({
    "#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!",
    "#GETTING_DATA", "#SPILL!", "#CALC!",
})

# El ajuste que obliga a recalcular. Se le pasa a LibreOffice en un perfil
# de usuario propio y desechable, para no tocarle la configuración al
# usuario ni chocar con una ventana de LibreOffice que esté abierta.
AJUSTE_RECALCULAR = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry"
 xmlns:xs="http://www.w3.org/2001/XMLSchema"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop
 oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
<item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop
 oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>
"""


def buscar_libreoffice():
    """Devuelve la ruta del programa de LibreOffice, o None si no está.

    Primero mira las rutas típicas de cada sistema y después busca en el
    PATH, por si el usuario lo instaló en otro lado.
    """
    for ruta in RUTAS_CONOCIDAS:
        candidata = Path(ruta)
        if candidata.exists():
            return candidata
    for nombre in ("soffice", "libreoffice"):
        encontrada = shutil.which(nombre)
        if encontrada:
            return Path(encontrada)
    return None


# ---------------------------------------------------------------------------
# Los valores ya calculados de un libro, leídos una sola vez
# ---------------------------------------------------------------------------
#
# Abrir el libro con openpyxl tarda casi un segundo: son 1,3 MB y 15 hojas,
# y openpyxl los vuelve a leer enteros cada vez. Antes, armar el formulario
# de un cliente abría el mismo archivo tres veces seguidas para preguntarle
# cosas distintas: si tenía celdas con error, cuáles eran los totales, y qué
# había en la hoja de captura. Ahora se lee una vez y las tres preguntas se
# contestan de la memoria.
#
# La llave es la ruta con la fecha y el tamaño del archivo. Si el archivo
# cambia —porque se volvió a generar o porque LibreOffice lo reescribió—,
# la llave cambia sola y se vuelve a leer. Nunca se devuelve algo viejo.
#
# El candado hace falta porque el servidor atiende varias peticiones a la
# vez, cada una en su hilo. Sin él, dos hilos pueden entrar juntos: uno
# vacía la memoria mientras el otro está leyendo de ella.

_calculados = {}
_candado_calculados = threading.Lock()

# Cuántos libros se recuerdan a la vez. Son 6.255 celdas por libro, unos
# pocos cientos de kilobytes: el techo es para no crecer sin final, no
# porque pese.
_CUANTOS_LIBROS = 4


def llave_de_archivo(ruta):
    """Identifica una versión concreta de un archivo: ruta, fecha y tamaño.

    Se usa st_mtime_ns (nanosegundos) y no st_mtime: en Windows la fecha
    en segundos puede repetirse entre dos escrituras seguidas, y entonces
    se devolvería el contenido viejo del archivo nuevo.
    """
    datos = ruta.stat()
    return (str(ruta.resolve()), datos.st_mtime_ns, datos.st_size)


def valores_calculados_del_libro(ruta_xlsx):
    """Todos los valores calculados del libro: {'hoja': {'G32': 1500000}}.

    Se abre con data_only=True, que trae los RESULTADOS de las fórmulas en
    vez de las fórmulas mismas. Esta carga es de solo lectura y NUNCA se
    guarda: si se guardara, borraría las 902 fórmulas del libro.

    Lo que se devuelve son valores sueltos, no el libro de openpyxl. Así
    nadie puede modificar sin querer lo que está guardado en la memoria.
    """
    ruta = Path(ruta_xlsx)
    llave = llave_de_archivo(ruta)

    with _candado_calculados:
        if llave in _calculados:
            return _calculados[llave]

        # read_only=True: openpyxl lee el archivo por partes en vez de
        # armar el libro entero en memoria. Aquí solo se recorre y se lee,
        # y tarda 3,5 veces menos: 0,24 segundos en vez de 0,85. Se
        # comprobó que devuelve exactamente las mismas 6.255 celdas.
        libro = load_workbook(ruta, data_only=True, read_only=True)
        try:
            leidos = {}
            for nombre in libro.sheetnames:
                hoja = {}
                for fila in libro[nombre].iter_rows():
                    for celda in fila:
                        if celda.value is not None:
                            hoja[celda.coordinate] = celda.value
                leidos[nombre] = hoja
        finally:
            # Obligatorio en este modo: deja el archivo abierto, y en
            # Windows un archivo abierto no se puede reemplazar. Sin esto,
            # LibreOffice no podría escribir encima el libro recalculado.
            libro.close()

        if len(_calculados) >= _CUANTOS_LIBROS:
            _calculados.clear()
        _calculados[llave] = leidos
        return leidos


def celdas_con_error(ruta_xlsx):
    """Lista las celdas que quedaron mostrando un error de Excel.

    Se mira sobre los valores calculados del libro, que es la única forma
    de ver un #REF!: las fórmulas no dicen si fallaron, solo sus resultados.
    """
    encontradas = []
    for nombre, hoja in valores_calculados_del_libro(ruta_xlsx).items():
        for coordenada, valor in hoja.items():
            if isinstance(valor, str) and valor.strip() in ERRORES_DE_EXCEL:
                encontradas.append(f"{nombre}!{coordenada}={valor.strip()}")
    return encontradas


def recalcular(ruta_xlsx, tiempo_limite=TIEMPO_LIMITE):
    """Recalcula el libro con LibreOffice y lo vuelve a guardar.

    Devuelve un informe (diccionario) y NUNCA lanza una excepción por que
    LibreOffice falte o falle: en ese caso el archivo se queda como estaba
    y el informe explica por qué. El programa tiene que seguir andando.

        {"recalculado": True/False,
         "motivo": "...",              texto para mostrarle al contador
         "segundos": 1.8,
         "programa": "/Applications/..."}
    """
    ruta_xlsx = Path(ruta_xlsx)
    informe = {
        "recalculado": False,
        "motivo": "",
        "segundos": 0.0,
        "programa": "",
    }

    programa = buscar_libreoffice()
    if programa is None:
        informe["motivo"] = (
            "LibreOffice no está instalado, así que no se pudieron calcular"
            " los totales. El archivo está completo y correcto: los totales"
            " aparecen al abrirlo en Excel."
        )
        return informe
    informe["programa"] = str(programa)

    comenzó = time.monotonic()

    # Carpetas desechables: un perfil de usuario propio para LibreOffice y
    # una carpeta de salida. Las dos se borran solas al terminar.
    with tempfile.TemporaryDirectory() as temporal:
        temporal = Path(temporal)
        perfil = temporal / "perfil"
        (perfil / "user").mkdir(parents=True)
        # encoding explícito: sin esto, en Windows se escribe en cp1252.
        (perfil / "user" / "registrymodifications.xcu").write_text(
            AJUSTE_RECALCULAR, encoding="utf-8"
        )
        salida = temporal / "salida"
        salida.mkdir()

        orden = [
            str(programa),
            "--headless",       # sin ventanas
            "--norestore",      # que no intente recuperar documentos
            "--nolockcheck",
            f"-env:UserInstallation={perfil.as_uri()}",
            "--convert-to", "xlsx",
            "--outdir", str(salida),
            str(ruta_xlsx),
        ]

        try:
            resultado = subprocess.run(
                orden,
                capture_output=True,
                timeout=tiempo_limite,
                # encoding explícito y errors="replace": los mensajes de
                # LibreOffice traen tildes, y en Windows se leen en cp1252
                # y revientan la lectura.
                text=True,
                encoding="utf-8",
                errors="replace",
            )
        except subprocess.TimeoutExpired:
            informe["segundos"] = round(time.monotonic() - comenzó, 1)
            informe["motivo"] = (
                f"LibreOffice se demoró más de {tiempo_limite} segundos y se"
                f" canceló. El archivo está completo: los totales aparecen"
                f" al abrirlo en Excel."
            )
            return informe
        except OSError as error:
            informe["segundos"] = round(time.monotonic() - comenzó, 1)
            informe["motivo"] = (
                f"No se pudo ejecutar LibreOffice ({error}). El archivo está"
                f" completo: los totales aparecen al abrirlo en Excel."
            )
            return informe

        informe["segundos"] = round(time.monotonic() - comenzó, 1)

        if resultado.returncode != 0:
            informe["motivo"] = (
                f"LibreOffice terminó con un error (código"
                f" {resultado.returncode}). El archivo está completo: los"
                f" totales aparecen al abrirlo en Excel."
            )
            return informe

        recalculado = salida / ruta_xlsx.name
        if not recalculado.exists():
            informe["motivo"] = (
                "LibreOffice no devolvió ningún archivo. El archivo está"
                " completo: los totales aparecen al abrirlo en Excel."
            )
            return informe

        # Solo aquí se reemplaza el archivo. Si algo hubiera fallado antes,
        # el archivo original de trabajo sigue intacto.
        shutil.copy2(recalculado, ruta_xlsx)

    informe["recalculado"] = True
    informe["motivo"] = (
        f"Totales calculados con LibreOffice en {informe['segundos']}"
        f" segundos."
    )
    return informe
