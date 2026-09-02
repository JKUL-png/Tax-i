"""
El lector del reporte de información exógena de la DIAN.

La exógena es lo que los terceros —bancos, empleadores, notarías,
municipios— le reportaron a la DIAN sobre un contribuyente durante el
año. El contador la descarga del portal de la DIAN como un archivo de
Excel y es lo primero que mira cuando arma una declaración: le dice qué
sabe la DIAN de su cliente.

Este archivo hace UNA cosa: abrir ese Excel y devolver lo que dice, en
datos que el resto del programa pueda usar. No guarda nada, no toca la
base de datos y no escribe archivos. Por eso se puede probar solo, sin
prender el servidor.

**Aquí no hay inteligencia artificial, y no la necesita.** El reporte es
una tabla bien formada: se lee con código y punto. Nada de lo que hay
aquí adentro sale del computador.

Tres cosas que este archivo NO hace, porque son la línea legal del
proyecto (ver CLAUDE.md):

  - No calcula impuestos, ni sumas, ni saldos.
  - No decide cuando la DIAN ofrece varias opciones. Cuando una fila
    dice «R29 si el saldo es positivo, R30 si es negativo», se marca
    como que requiere decisión y se le muestran al contador las dos
    opciones tal como la DIAN las escribió. Elegir es criterio
    profesional.
  - No une ni descarta posibles duplicados. Solo los marca y dice por
    qué se marcaron.

Los tres avisos legales de la cabecera se guardan **textuales**. Son de
la DIAN, no nuestros, y el tercero es el que dice que la exógena no
reemplaza la realidad económica del contribuyente.
"""

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


class ExogenaInvalida(ValueError):
    """El archivo no es un reporte de exógena que se pueda leer.

    Lleva un mensaje escrito para que lo entienda el contador, no para
    que lo entienda un programador.
    """


# ----------------------------------------------------------
# Cómo se reconocen las columnas
# ----------------------------------------------------------
#
# NO se usan números de fila fijos. La fila de encabezados se busca por
# el texto «Uso declaración Sugerida», y todo lo demás se ubica a partir
# de ahí. Si la DIAN mueve la tabla el año que viene, esto sigue
# funcionando; si le cambia el nombre a esa columna, el programa avisa
# claro en vez de leer cualquier cosa.
ENCABEZADO_USO = "uso declaracion sugerida"
ENCABEZADO_DETALLE = "detalle"
ENCABEZADO_VALOR = "valor"
ENCABEZADO_NIT = "nit"
ENCABEZADO_ADICIONAL = "informacion adicional"

# El código de concepto viene metido dentro del detalle, así:
#     «Pagos por salarios (Concepto: 2276)»
CONCEPTO = re.compile(r"concepto\s*:\s*(\d+)", re.IGNORECASE)

# Los renglones del formulario 210 aparecen dentro del uso sugerido con
# la forma R seguida de números: R29, R132. El \b de la derecha importa:
# sin él, «R100» se leería como «R10».
CODIGO_RENGLON = re.compile(r"\bR(\d+)\b")

# Los topes se citan de dos maneras distintas en el mismo archivo:
# «Tope 1 - Ingresos» en las filas de resumen y «Tope 1: Ingresos
# brutos» dentro del uso sugerido. Por eso se enlazan por el NÚMERO y
# nunca por el nombre: los nombres no coinciden.
TOPE = re.compile(r"tope\s*(\d+)\s*[-:]\s*([^|\n]*)", re.IGNORECASE)

# La DIAN separa una nota aclaratoria del resto del uso sugerido con un
# salto de línea y la palabra «Nota:». Se guarda aparte para que no se
# confunda con las opciones que el contador tiene que elegir.
MARCA_DE_NOTA = re.compile(r"\n\s*nota\s*:", re.IGNORECASE)

# Un texto largo en la primera columna, antes de la tabla, es uno de los
# avisos legales. Las etiquetas de la cabecera («Tipo de documento:»,
# «Nombres / Razón social:») son todas mucho más cortas que esto.
LARGO_MINIMO_DE_AVISO = 60

# Dos cifras se consideran parecidas cuando se diferencian en menos de
# esto. Es el umbral que pidió el contador: 1%.
DIFERENCIA_MAXIMA = 0.01


# ----------------------------------------------------------
# Herramientas pequeñas
# ----------------------------------------------------------


def _normalizar(texto):
    """Deja un texto comparable: sin tildes, sin mayúsculas, sin espacios de más.

    Hace falta porque el archivo real trae cosas como
    «Información  Adicional » —con doble espacio y espacio al final— y
    porque la DIAN podría quitarle la tilde a «declaración» algún año.
    """
    if texto is None:
        return ""
    limpio = " ".join(str(texto).split()).lower()
    sin_tildes = unicodedata.normalize("NFD", limpio)
    return "".join(c for c in sin_tildes if unicodedata.category(c) != "Mn")


def _texto(valor):
    """El contenido de una celda como texto, sin espacios sobrantes afuera.

    Los espacios de ADENTRO se respetan: la DIAN escribe «R30 Deudas  (si
    el saldo es negativo)» con dos espacios, y cuando se le muestran las
    opciones al contador van palabra por palabra, tal como ella las
    escribió.
    """
    if valor is None:
        return ""
    return str(valor).strip()


def _identificacion(valor):
    """Un NIT o una cédula, siempre como texto.

    En el archivo vienen mezclados: el NIT de quien reporta llega como
    número y el del contribuyente como texto. Si se dejaran como número
    se perderían los ceros de la izquierda y aparecería notación
    científica en pantalla. Así que todos terminan en texto, sin
    decimales.
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


def _fecha(valor):
    """Una fecha del Excel en formato ISO, o texto tal cual si no es fecha."""
    if valor is None:
        return ""
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    return str(valor).strip()


def _numero(valor):
    """El valor de una celda como número. Devuelve None si no es número.

    No redondea ni convierte de moneda: devuelve lo que hay.
    """
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return valor
    return None


def _segmentos(texto):
    """Parte el uso sugerido por donde la DIAN lo separó: la barra vertical."""
    return [parte.strip() for parte in texto.split("|") if parte.strip()]


# ----------------------------------------------------------
# Encontrar la tabla dentro de la hoja
# ----------------------------------------------------------


def _abrir_hoja(ruta):
    """Abre el libro y devuelve la hoja del reporte."""
    ruta = Path(ruta)
    if not ruta.exists():
        raise ExogenaInvalida("No se encuentra el archivo %s." % ruta.name)
    try:
        libro = load_workbook(ruta, data_only=True)
    except Exception:
        raise ExogenaInvalida(
            "No se pudo abrir «%s» como archivo de Excel. La exógena se"
            " descarga del portal de la DIAN en formato .xlsx." % ruta.name
        )
    if "Reporte" in libro.sheetnames:
        return libro["Reporte"]
    # Si algún año le cambian el nombre a la hoja y solo hay una, se usa
    # esa. Con varias no se adivina: se avisa.
    if len(libro.sheetnames) == 1:
        return libro[libro.sheetnames[0]]
    raise ExogenaInvalida(
        "El archivo no tiene una hoja llamada «Reporte». Tiene: %s."
        % ", ".join(libro.sheetnames)
    )


def _buscar_encabezado(hoja):
    """Encuentra la fila de encabezados buscando «Uso declaración Sugerida».

    Devuelve el número de esa fila. Es el ancla de todo lo demás.
    """
    for fila in range(1, hoja.max_row + 1):
        for columna in range(1, hoja.max_column + 1):
            if _normalizar(hoja.cell(row=fila, column=columna).value) == ENCABEZADO_USO:
                return fila
    raise ExogenaInvalida(
        "El archivo no tiene la columna «Uso declaración Sugerida», así que"
        " no parece el reporte de información exógena que se descarga del"
        " portal de la DIAN. Revise que no sea otro archivo."
    )


def _ubicar_columnas(hoja, fila_encabezado):
    """Dice en qué columna está cada cosa, leyendo los encabezados.

    Dos columnas se llaman «NIT» y dos empiezan por «Nombre»: la primera
    de cada par es la de quien reporta y la segunda la del
    contribuyente. Se distinguen por el orden, que es como viene el
    archivo, no por el texto.
    """
    nits = []
    nombres = []
    columnas = {}

    for columna in range(1, hoja.max_column + 1):
        titulo = _normalizar(hoja.cell(row=fila_encabezado, column=columna).value)
        if not titulo:
            continue
        if titulo == ENCABEZADO_NIT:
            nits.append(columna)
        elif titulo.startswith("nombre"):
            nombres.append(columna)
        elif titulo == ENCABEZADO_DETALLE:
            columnas["detalle"] = columna
        elif titulo == ENCABEZADO_VALOR:
            columnas["valor"] = columna
        elif titulo == ENCABEZADO_USO:
            columnas["uso"] = columna
        elif titulo == ENCABEZADO_ADICIONAL:
            columnas["adicional"] = columna

    if nits:
        columnas["nit_reporta"] = nits[0]
    if len(nits) > 1:
        columnas["nit_contribuyente"] = nits[1]
    if nombres:
        columnas["nombre_reporta"] = nombres[0]
    if len(nombres) > 1:
        columnas["nombre_contribuyente"] = nombres[1]

    faltantes = [
        nombre for nombre in ("detalle", "valor", "uso") if nombre not in columnas
    ]
    if faltantes:
        traduccion = {"detalle": "Detalle", "valor": "Valor",
                      "uso": "Uso declaración Sugerida"}
        raise ExogenaInvalida(
            "A la tabla del archivo le faltan columnas: %s."
            % ", ".join("«%s»" % traduccion[f] for f in faltantes)
        )
    return columnas


# ----------------------------------------------------------
# La cabecera: de quién es el reporte y los avisos de la DIAN
# ----------------------------------------------------------

# Las etiquetas de la cabecera, ya normalizadas, y el nombre con el que
# se devuelve cada una. Ojo: «identificacion» e «identificacion del
# consultante» son cosas distintas, por eso la comparación es exacta.
ETIQUETAS = {
    "ano al que se refiere la consulta": "anio",
    "fecha corte del proceso": "fecha_corte",
    "tipo de documento": "tipo_documento",
    "identificacion": "identificacion",
    "nombres / razon social": "nombre",
}


def _leer_cabecera(hoja, fila_encabezado):
    """Lee de quién es el reporte y cuándo se sacó.

    Busca por la etiqueta escrita en la primera columna, no por número
    de fila. El valor es la primera celda con contenido a la derecha,
    porque la DIAN a veces la pone en la columna B y a veces en la C.
    """
    cabecera = {
        "anio": "", "fecha_corte": "", "fecha_reporte": "",
        "tipo_documento": "", "identificacion": "", "nombre": "",
    }

    for fila in range(1, fila_encabezado):
        for columna in range(1, hoja.max_column + 1):
            etiqueta = _normalizar(hoja.cell(row=fila, column=columna).value)
            etiqueta = etiqueta.rstrip(":").strip()

            # La fecha en que se generó el reporte, que no es lo mismo
            # que la fecha de corte: el corte es hasta cuándo alcanzaron
            # a llegar los datos.
            if etiqueta in ("fecha reporte", "fecha de reporte"):
                cabecera["fecha_reporte"] = _fecha(
                    _primer_valor_a_la_derecha(hoja, fila, columna)
                )
                continue

            if etiqueta not in ETIQUETAS:
                continue
            crudo = _primer_valor_a_la_derecha(hoja, fila, columna)
            campo = ETIQUETAS[etiqueta]
            if campo == "fecha_corte":
                cabecera[campo] = _fecha(crudo)
            elif campo == "identificacion":
                cabecera[campo] = _identificacion(crudo)
            else:
                cabecera[campo] = _texto(crudo)

    return cabecera


def _primer_valor_a_la_derecha(hoja, fila, columna):
    """El primer contenido que haya después de una etiqueta, en su misma fila."""
    for siguiente in range(columna + 1, hoja.max_column + 1):
        valor = hoja.cell(row=fila, column=siguiente).value
        if valor is not None and str(valor).strip():
            return valor
    return None


def _leer_avisos(hoja, fila_encabezado):
    """Saca los avisos legales de la cabecera, TEXTUALES.

    No se resumen, no se reescriben y no se les corrige la puntuación.
    Son de la DIAN. El tercero es el que dice que la exógena no es
    indispensable y no reemplaza la realidad económica del
    contribuyente.

    Se reconocen porque son los textos largos de la primera columna: las
    etiquetas de la cabecera son todas cortas y terminan en dos puntos.
    """
    avisos = []
    for fila in range(1, fila_encabezado):
        valor = hoja.cell(row=fila, column=1).value
        if not isinstance(valor, str):
            continue
        limpio = valor.strip()
        if len(limpio) < LARGO_MINIMO_DE_AVISO or limpio.endswith(":"):
            continue
        avisos.append(valor)
    return avisos


# ----------------------------------------------------------
# Los cinco topes
# ----------------------------------------------------------


def _es_tope(detalle):
    """Si el texto es un tope, devuelve (número, nombre). Si no, None."""
    encontrado = re.match(r"\s*tope\s*(\d+)\s*[-:]\s*(.*)", detalle, re.IGNORECASE)
    if not encontrado:
        return None
    return int(encontrado.group(1)), encontrado.group(2).strip()


# ----------------------------------------------------------
# Leer una fila de datos
# ----------------------------------------------------------


def _partir_uso(uso):
    """Separa el uso sugerido en la parte de las opciones y la nota final.

    La DIAN cierra algunas filas con «\\nNota: este valor puede ser
    reportado por el empleador y el fondo de cesantías...». Eso es una
    aclaración, no una opción para elegir, así que va a un campo aparte.
    """
    marca = MARCA_DE_NOTA.search(uso)
    if not marca:
        return uso, ""
    return uso[: marca.start()], uso[marca.start():].strip()


def _renglones_de(uso):
    """Los renglones del 210 que menciona el uso sugerido.

    De cada uno se guarda el código y el PEDAZO LITERAL donde apareció,
    sin recortar. Ese pedazo es lo que ve el contador: la DIAN a veces
    escribe un nombre limpio («R30 Deudas») y a veces una frase entera
    que menciona varios códigos, y en los dos casos se muestra completa.
    """
    encontrados = []
    vistos = set()
    for segmento in _segmentos(uso):
        for hallazgo in CODIGO_RENGLON.finditer(segmento):
            codigo = "R" + hallazgo.group(1)
            if codigo in vistos:
                continue
            vistos.add(codigo)
            encontrados.append({
                "codigo": codigo,
                # Lo que sigue al código dentro de ese pedazo. Sirve para
                # deducir el nombre corto del renglón (ver más abajo).
                "cola": segmento[hallazgo.end():].strip(" .,:;"),
                # El pedazo completo, palabra por palabra.
                "texto": segmento,
            })
    return encontrados


def _topes_de(uso):
    """Los topes que menciona el uso sugerido, por número."""
    encontrados = []
    vistos = set()
    for hallazgo in TOPE.finditer(uso):
        numero = int(hallazgo.group(1))
        if numero in vistos:
            continue
        vistos.add(numero)
        encontrados.append({
            "numero": numero,
            "texto": hallazgo.group(0).strip(),
        })
    return encontrados


def _informacion_adicional(texto):
    """Convierte «Clave: Valor | Clave: Valor» en un diccionario.

    Los pedazos que no traigan dos puntos se ignoran en vez de romper la
    lectura: más vale perder un dato suelto que no poder abrir el
    archivo.
    """
    datos = {}
    if not texto:
        return datos
    for parte in texto.split("|"):
        parte = parte.strip()
        if not parte or ":" not in parte:
            continue
        clave, valor = parte.split(":", 1)
        clave = clave.strip()
        if clave:
            datos[clave] = valor.strip()
    return datos


# ----------------------------------------------------------
# El nombre corto de cada renglón
# ----------------------------------------------------------


def _nombres_de_renglones(filas):
    """Le pone a cada código de renglón el nombre que la DIAN le da.

    El problema: la DIAN no escribe el nombre igual en todas partes. En
    una fila dice «R29 Patrimonio Bruto» y en otra «R29 Patrimonio Bruto
    (si el saldo es positivo)». En otra escribe una frase corrida que se
    traga varios códigos: «R36 Otras rentas exentas (laborales) o Otras
    deducciones imputables para R51 (Honorarios)...». Y a veces el
    código solo viene con una etiqueta entre paréntesis: «R33 (Trabajo)».

    La solución no inventa nada: junta todo lo que el archivo dice de
    cada código, en todas las filas, y escoge así:

      1. Descarta los textos que se tragaron OTRO código: son frases,
         no nombres.
      2. Entre los que quedan, se queda con el más corto. Así «R29
         Patrimonio Bruto» le gana a «R29 Patrimonio Bruto (si el saldo
         es positivo)».
      3. Si no quedó ninguno, usa la etiqueta que la DIAN puso entre
         paréntesis: R33 queda como «Trabajo».

    Sobre el archivo de ejemplo esto le pone nombre a los quince
    códigos. Y el contador puede renombrar cualquiera, porque los
    renglones son suyos.
    """
    candidatos = {}
    for fila in filas:
        for renglon in fila["renglones"]:
            candidatos.setdefault(renglon["codigo"], []).append(renglon["cola"])

    nombres = {}
    for codigo, colas in candidatos.items():
        nombres[codigo] = _mejor_nombre(colas)
    return nombres


def _mejor_nombre(colas):
    """Escoge el nombre de un renglón entre todo lo que el archivo dice de él."""
    propios = []
    entre_parentesis = []
    for cola in colas:
        if not cola:
            continue
        if cola.startswith("("):
            etiqueta = re.match(r"\(([^)]+)\)", cola)
            if etiqueta:
                entre_parentesis.append(etiqueta.group(1).strip())
            continue
        if CODIGO_RENGLON.search(cola):
            # Se tragó otro código: es una frase, no un nombre.
            continue
        propios.append(cola)

    if propios:
        return min(propios, key=len)
    if entre_parentesis:
        return min(entre_parentesis, key=len)
    return ""


def catalogo_de_renglones(lectura):
    """Los renglones que hay que crearle al cliente, en orden de código.

    Cada uno trae el código, el nombre corto y el título ya armado tal
    como se ve en pantalla: «R32 — Ingresos brutos por rentas de
    trabajo (art. 103 E.T.)».
    """
    nombres = _nombres_de_renglones(lectura["filas"])
    catalogo = []
    for codigo in sorted(nombres, key=lambda c: int(c[1:])):
        nombre = nombres[codigo]
        catalogo.append({
            "codigo": codigo,
            "nombre": nombre,
            "titulo": "%s — %s" % (codigo, nombre) if nombre else codigo,
        })
    return catalogo


# ----------------------------------------------------------
# Posibles duplicados
# ----------------------------------------------------------


def _parecidos(uno, otro):
    """¿Son dos cifras iguales, o casi? «Casi» es menos del 1% de diferencia."""
    if uno is None or otro is None:
        return False
    if uno == otro:
        return True
    mayor = max(abs(uno), abs(otro))
    if mayor == 0:
        return False
    return abs(uno - otro) / mayor < DIFERENCIA_MAXIMA


def _marcar_duplicados(filas):
    """Marca las filas donde el mismo hecho económico pudo llegar dos veces.

    Dos terceros distintos pueden reportar lo mismo —la DIAN lo advierte
    ella misma con las cesantías, que las reporta el empleador y también
    el fondo— y a veces es el mismo tercero el que informa dos veces.

    Se marca cuando dos filas tienen un valor igual o casi igual Y
    apuntan al mismo renglón o al mismo tope. Cada marca dice POR QUÉ se
    marcó, para que el contador la descarte de un vistazo cuando no lo
    sea: no es lo mismo que dos terceros distintos coincidan a que el
    mismo tercero informe dos conceptos que valen igual.

    **Solo marca.** No une las filas, no descarta ninguna y no elige
    cuál vale. Eso es criterio profesional.
    """
    parejas = []
    for i, una in enumerate(filas):
        for otra in filas[i + 1:]:
            if not _parecidos(una["valor"], otra["valor"]):
                continue

            renglones_una = {r["codigo"] for r in una["renglones"]}
            renglones_otra = {r["codigo"] for r in otra["renglones"]}
            comunes = renglones_una & renglones_otra

            topes_una = {t["numero"] for t in una["topes"]}
            topes_otra = {t["numero"] for t in otra["topes"]}
            topes_comunes = topes_una & topes_otra

            if not comunes and not topes_comunes:
                continue

            if comunes:
                donde = "al mismo renglón (%s)" % ", ".join(sorted(comunes))
            else:
                donde = "al mismo tope (%s)" % ", ".join(
                    "Tope %d" % n for n in sorted(topes_comunes)
                )

            if una["nit_reporta"] != otra["nit_reporta"]:
                confianza = "alta"
                motivo = ("Otro tercero reporta un valor casi igual y apunta %s."
                          % donde)
            else:
                confianza = "media"
                motivo = ("El mismo tercero reporta un valor casi igual en dos"
                          " conceptos distintos, y los dos apuntan %s." % donde)

            pareja = {
                "filas": [una["fila_excel"], otra["fila_excel"]],
                "confianza": confianza,
                "motivo": motivo,
            }
            parejas.append(pareja)
            for cual, con_quien in ((una, otra), (otra, una)):
                cual["posible_duplicado"] = True
                cual["duplicado_de"].append({
                    "fila_excel": con_quien["fila_excel"],
                    "quien_reporta": con_quien["nombre_reporta"],
                    "detalle": con_quien["detalle"],
                    "valor": con_quien["valor"],
                    "confianza": confianza,
                    "motivo": motivo,
                })
    return parejas


# ----------------------------------------------------------
# La puerta de entrada
# ----------------------------------------------------------


def leer(ruta):
    """Lee el reporte de exógena y devuelve todo lo que dice.

    Devuelve un diccionario con:

      cabecera   — año, tipo y número de documento, nombre, fecha de
                   corte y fecha del reporte
      avisos     — los avisos legales de la DIAN, textuales
      topes      — los cinco topes, con su número, su nombre y su valor
      filas      — un registro por cada cosa que le reportaron
      renglones  — el catálogo de renglones del 210 que hay que crear
      duplicados — las parejas de filas que pueden ser el mismo hecho

    No guarda nada ni modifica el archivo: lo abre, lo lee y lo cierra.
    """
    hoja = _abrir_hoja(ruta)
    fila_encabezado = _buscar_encabezado(hoja)
    columnas = _ubicar_columnas(hoja, fila_encabezado)

    def celda(fila, cual):
        if cual not in columnas:
            return None
        return hoja.cell(row=fila, column=columnas[cual]).value

    topes = []
    filas = []

    for numero_fila in range(fila_encabezado + 1, hoja.max_row + 1):
        detalle = _texto(celda(numero_fila, "detalle"))
        valor = _numero(celda(numero_fila, "valor"))
        nit_reporta = _identificacion(celda(numero_fila, "nit_reporta"))

        if not detalle and valor is None and not nit_reporta:
            continue

        # Los cinco topes son resumen, no renglones: van aparte. Se
        # reconocen porque dicen «Tope N -» y no tienen reportante.
        tope = _es_tope(detalle) if not nit_reporta else None
        if tope:
            numero, nombre = tope
            topes.append({
                "numero": numero,
                "nombre": nombre,
                "etiqueta": detalle,
                "valor": valor,
                "fila_excel": numero_fila,
            })
            continue

        uso = _texto(celda(numero_fila, "uso"))
        opciones_texto, nota = _partir_uso(uso)
        renglones = _renglones_de(opciones_texto)
        concepto = CONCEPTO.search(detalle)

        # Cuando la DIAN propone más de un renglón para la misma cifra,
        # la fila requiere una decisión del contador. Las opciones se
        # le muestran tal como ella las escribió, palabra por palabra.
        # Tax-i nunca elige: ni con IA, ni con reglas, ni mirando el
        # signo del valor.
        opciones = [
            segmento for segmento in _segmentos(opciones_texto)
            if CODIGO_RENGLON.search(segmento)
        ]
        requiere_decision = len(renglones) > 1

        filas.append({
            "fila_excel": numero_fila,
            "nit_reporta": nit_reporta,
            "nombre_reporta": _texto(celda(numero_fila, "nombre_reporta")),
            "nit_contribuyente": _identificacion(
                celda(numero_fila, "nit_contribuyente")
            ),
            "nombre_contribuyente": _texto(
                celda(numero_fila, "nombre_contribuyente")
            ),
            "detalle": detalle,
            "concepto": concepto.group(1) if concepto else "",
            "valor": valor,
            "uso_sugerido": uso,
            "nota": nota,
            "renglones": renglones,
            "opciones": opciones,
            "requiere_decision": requiere_decision,
            "topes": _topes_de(uso),
            "informacion_adicional": _informacion_adicional(
                _texto(celda(numero_fila, "adicional"))
            ),
            "posible_duplicado": False,
            "duplicado_de": [],
        })

    duplicados = _marcar_duplicados(filas)

    lectura = {
        "hoja": hoja.title,
        "fila_encabezado": fila_encabezado,
        "cabecera": _leer_cabecera(hoja, fila_encabezado),
        "avisos": _leer_avisos(hoja, fila_encabezado),
        "topes": topes,
        "filas": filas,
        "duplicados": duplicados,
    }
    lectura["renglones"] = catalogo_de_renglones(lectura)
    return lectura


def resumen(lectura):
    """Un conteo corto de lo que trajo el archivo, para mostrar al cargarlo."""
    return {
        "anio": lectura["cabecera"]["anio"],
        "nombre": lectura["cabecera"]["nombre"],
        "identificacion": lectura["cabecera"]["identificacion"],
        "fecha_corte": lectura["cabecera"]["fecha_corte"],
        "registros": len(lectura["filas"]),
        "topes": len(lectura["topes"]),
        "renglones": len(lectura["renglones"]),
        "requieren_decision": sum(
            1 for f in lectura["filas"] if f["requiere_decision"]
        ),
        "posibles_duplicados": sum(
            1 for f in lectura["filas"] if f["posible_duplicado"]
        ),
        "avisos": len(lectura["avisos"]),
    }
