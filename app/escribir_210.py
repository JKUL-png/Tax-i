"""
Escritura de valores en la copia de la plantilla del Formulario 210.

Este módulo es el único del programa que escribe dentro de un archivo de
Excel, y está construido para que sea difícil hacer daño. Las reglas están
en el código, no en un comentario:

  1. Antes de escribir en una celda se verifica que no tenga fórmula. Si la
     tiene, se lanza una excepción y no se escribe nada. No hay bandera ni
     parámetro para saltarse esta verificación.
  2. Nunca se abre el libro con data_only=True para guardarlo. La única
     carga que existe aquí es de solo lectura y jamás se guarda: si se
     guardara, reemplazaría las 902 fórmulas del libro por números y las
     destruiría para siempre.
  3. Nunca se escribe sobre el archivo original. Lo primero que se hace es
     copiarlo a la carpeta de trabajo. El archivo del contador queda igual
     pase lo que pase.
  4. Solo se escribe en la hoja «Detalle renglón 210». Cualquier intento de
     escribir en otra hoja se rechaza.
  5. Cada escritura queda en una bitácora que se guarda junto al archivo de
     salida, para que el contador pueda revisar y revertir.
  6. Después de guardar, el archivo se vuelve a abrir y se comparan TODAS
     sus fórmulas contra las del original. Si cambió aunque sea una, el
     resultado se borra y se avisa. No se entrega un archivo dudoso.

Cómo escribe (y por qué así)
----------------------------
Un .xlsx es un ZIP con archivos XML adentro. La forma fácil de escribir
sería abrir con openpyxl y guardar, pero openpyxl reescribe el archivo
completo y bota lo que no entiende: en esta plantilla se pierden 22
imágenes, entre ellas las tres de la hoja Copyright. Es una plantilla
comercial con licencia de un tercero, así que eso no es aceptable.

Entonces se hace la escritura "quirúrgica": se abre el ZIP, se copian TODAS
las partes tal cual, byte por byte, y solo se toca el XML de la hoja de
captura, cambiando únicamente el valor de las celdas pedidas. No se pierde
nada: ni imágenes, ni notas, ni formatos, ni anchos de columna.
"""

import json
import re
import shutil
import threading
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

from app.documentos import sanitizar_nombre
from app.recalcular import celdas_con_error, llave_de_archivo, recalcular
from app.plantilla_210 import (
    COLUMNAS_VALOR,
    FILA_INICIAL,
    HOJA_CAPTURA,
    RAIZ,
    _mapa_de_combinadas,
    _ultima_fila_con_contenido,
    es_formula,
)

# Carpeta donde viven las copias de trabajo. Está dentro de datos/, que no
# se sube a git.
CARPETA_TRABAJO = RAIZ / "datos" / "trabajo"

# Terminación del archivo de bitácora que acompaña a cada salida.
SUFIJO_BITACORA = ".bitacora.json"


class VerificacionFallida(Exception):
    """El archivo generado no pasó la revisión posterior.

    Es un error grave: quiere decir que al guardar se dañó algo del libro.
    Cuando esto pasa, el archivo generado se borra. Es preferible quedarse
    sin archivo que entregarle al contador uno con las fórmulas rotas.
    """


class EscrituraBloqueada(Exception):
    """Se pidió una escritura que las reglas no permiten.

    Es a propósito una excepción propia: el resto del programa la puede
    atrapar y mostrarle al contador qué se rechazó y por qué, sin
    confundirla con un error cualquiera de Python.
    """


def _formatear_numero(valor):
    """Convierte el número a texto tal como Excel lo guarda en el XML.

    Sin separadores de miles, sin símbolo de peso y con punto decimal. El
    formato con el que se ve en pantalla lo pone el estilo de la celda, que
    no se toca.
    """
    if isinstance(valor, int):
        return str(valor)
    if float(valor).is_integer():
        return str(int(valor))
    return repr(float(valor))


def preparar_copia(ruta_plantilla, nombre_salida=None, carpeta=CARPETA_TRABAJO):
    """Copia la plantilla a la carpeta de trabajo y devuelve la ruta de la copia.

    Regla 3: el original nunca se toca. Todo lo que sigue pasa sobre esta
    copia. shutil.copy2 conserva la fecha del archivo.
    """
    origen = Path(ruta_plantilla)
    if not origen.exists():
        raise FileNotFoundError(f"No se encontró la plantilla: {origen}")

    if nombre_salida is None:
        nombre_salida = f"{origen.stem}_diligenciado.xlsx"
    # Sanitizado porque el nombre puede venir del nombre de un cliente, y
    # en Windows un ':' o un '?' rompen el guardado.
    nombre_salida = sanitizar_nombre(nombre_salida)
    if not nombre_salida.lower().endswith(".xlsx"):
        nombre_salida += ".xlsx"

    carpeta = Path(carpeta)
    carpeta.mkdir(parents=True, exist_ok=True)
    destino = carpeta / nombre_salida

    if destino.resolve() == origen.resolve():
        raise EscrituraBloqueada(
            "La copia de trabajo apunta al mismo archivo que la plantilla"
            " original. Se cancela para no escribir sobre el original."
        )

    shutil.copy2(origen, destino)
    return destino


# ---------------------------------------------------------------------------
# La ficha de la plantilla: lo que hay que saber de ella para escribir
# ---------------------------------------------------------------------------
#
# Antes, cada vez que se armaba el formulario de un cliente se abría el
# libro entero con openpyxl solo para preguntarle cuatro cosas. Abrirlo
# tarda casi un segundo (1,3 MB, 15 hojas), y las cuatro respuestas son
# siempre las mismas mientras la plantilla no cambie.
#
# Ahora se leen una vez y se guardan. Se lee de la plantilla ORIGINAL y no
# de la copia de trabajo, y eso es correcto porque `preparar_copia` acaba
# de copiar el original tal cual: en ese momento los dos archivos son
# idénticos byte por byte.

_fichas = {}
_candado_fichas = threading.Lock()
_CUANTAS_FICHAS = 3


def _ficha_de_plantilla(ruta_plantilla):
    """Lo que hace falta saber del libro para poder escribir en él.

    Devuelve un diccionario con:
      - 'anclas'      : para cada celda combinada, dónde vive de verdad
      - 'fila_final'  : hasta qué fila llega la tabla
      - 'valores'     : lo que hay escrito en cada celda de la hoja de
                        captura, incluidas las fórmulas tal como están

    Todo son datos sueltos, no el libro de openpyxl: así nadie puede
    modificar sin querer lo que quedó guardado en memoria.
    """
    ruta = Path(ruta_plantilla)
    llave = llave_de_archivo(ruta)

    with _candado_fichas:
        if llave in _fichas:
            return _fichas[llave]

        libro = load_workbook(ruta, data_only=False)
        if HOJA_CAPTURA not in libro.sheetnames:
            raise EscrituraBloqueada(
                f"La plantilla no tiene la hoja «{HOJA_CAPTURA}»."
            )
        hoja = libro[HOJA_CAPTURA]
        ficha = {
            "anclas": _mapa_de_combinadas(hoja),
            "fila_final": _ultima_fila_con_contenido(hoja),
            "valores": {
                celda.coordinate: celda.value
                for fila in hoja.iter_rows()
                for celda in fila
                if celda.value is not None
            },
        }

        if len(_fichas) >= _CUANTAS_FICHAS:
            _fichas.clear()
        _fichas[llave] = ficha
        return ficha


class EscritorPlantilla:
    """Escribe valores en la copia de trabajo de una plantilla.

    Se usa así:

        escritor = EscritorPlantilla("plantillas/mi_plantilla.xlsx")
        escritor.escribir("G32", 1500000, documento="certificado_banco.pdf")
        escritor.escribir("H104", 0, documento="revisión manual")
        ruta, bitacora = escritor.guardar()

    Las escrituras se validan en el momento en que se piden, pero el
    archivo solo se modifica al llamar guardar(). Así, si una sola de ellas
    está mal, no queda un archivo a medio llenar.
    """

    def __init__(self, ruta_plantilla, nombre_salida=None,
                 carpeta_trabajo=CARPETA_TRABAJO):
        self.ruta_original = Path(ruta_plantilla)
        self.ruta_copia = preparar_copia(
            self.ruta_original, nombre_salida, carpeta_trabajo
        )

        # Lo que hay que saber del libro para escribir en él: qué hay en
        # cada celda, cuáles tienen fórmula, cuáles están combinadas y
        # hasta dónde llega la tabla. Se consulta, nunca se guarda.
        #
        # Se saca de la plantilla original y no de la copia porque acaban
        # de ser el mismo archivo, y así la lectura sirve para todos los
        # clientes en vez de repetirse en cada uno.
        ficha = _ficha_de_plantilla(self.ruta_original)
        self._anclas = ficha["anclas"]
        self._fila_final = ficha["fila_final"]
        self._valores = ficha["valores"]

        # Cambios pendientes: coordenada -> valor nuevo.
        self.cambios = {}
        self.bitacora = []
        self.guardado = False
        # Se llenan al guardar, con el resultado de la verificación
        # posterior y del recálculo.
        self.informe_verificacion = None
        self.informe_recalculo = None

    # -- validación ---------------------------------------------------

    def _revisar_hoja(self, hoja):
        """Regla 4: solo se escribe en la hoja de captura."""
        if hoja != HOJA_CAPTURA:
            raise EscrituraBloqueada(
                f"Escritura rechazada: solo se puede escribir en la hoja"
                f" «{HOJA_CAPTURA}», y se pidió «{hoja}»."
            )

    def _revisar_celda(self, celda):
        """Que la coordenada exista, esté en el rango y en una columna de valores."""
        try:
            fila, columna = coordinate_to_tuple(celda)
        except Exception:
            raise EscrituraBloqueada(
                f"Escritura rechazada: «{celda}» no es una celda válida."
            )

        # get_column_letter y no hoja.cell(...).column_letter: si la celda
        # es parte de un grupo combinado, openpyxl devuelve un objeto
        # distinto que no tiene ese atributo y revienta.
        letra = get_column_letter(columna)
        if letra not in COLUMNAS_VALOR:
            raise EscrituraBloqueada(
                f"Escritura rechazada: {celda} está en la columna {letra}, y"
                f" solo se escribe en las columnas de valores"
                f" ({', '.join(COLUMNAS_VALOR)})."
            )
        if not (FILA_INICIAL <= fila <= self._fila_final):
            raise EscrituraBloqueada(
                f"Escritura rechazada: la fila {fila} está fuera de la tabla"
                f" (va de la {FILA_INICIAL} a la {self._fila_final})."
            )
        return fila, letra

    def _revisar_valor(self, valor):
        """Solo números.

        En estas columnas van cifras de dinero. Un texto obligaría a tocar
        la tabla de textos compartidos del archivo, que es justo lo que no
        se quiere. Y para limpiar una casilla se escribe 0, no vacío: es la
        convención del propio archivo.
        """
        if isinstance(valor, bool) or not isinstance(valor, (int, float)):
            raise EscrituraBloqueada(
                f"Escritura rechazada: solo se escriben números, y llegó"
                f" {type(valor).__name__} ({valor!r})."
                f" Para limpiar una casilla se escribe 0, no vacío."
            )

    def _revisar_formula(self, celda):
        """Regla 1: la verificación que protege el archivo.

        Si la celda tiene una fórmula, se aborta. Sin excepciones y sin
        forma de saltárselo.
        """
        valor = self._valores.get(celda)
        if es_formula(valor):
            raise EscrituraBloqueada(
                f"{HOJA_CAPTURA}!{celda} contiene una fórmula — escritura"
                f" bloqueada: {valor}"
            )

    def _revisar_combinada(self, celda):
        """Una celda combinada solo guarda el valor en su esquina superior izquierda."""
        ancla = self._anclas.get(celda)
        if ancla is not None and ancla != celda:
            raise EscrituraBloqueada(
                f"Escritura rechazada: {celda} es parte de un grupo de celdas"
                f" combinadas. El valor iría en {ancla}."
            )

    # -- escritura ----------------------------------------------------

    def escribir(self, celda, valor, documento, hoja=HOJA_CAPTURA):
        """Anota una escritura, después de pasar todas las verificaciones.

        'documento' es de dónde salió el dato: el nombre del archivo que lo
        respalda, o una nota como 'digitado por el contador'. Queda en la
        bitácora para que después se pueda rastrear.
        """
        if self.guardado:
            raise EscrituraBloqueada(
                "Este archivo ya se guardó. Para más cambios hay que empezar"
                " otra escritura."
            )

        celda = str(celda).upper().strip()
        self._revisar_hoja(hoja)
        self._revisar_celda(celda)
        self._revisar_valor(valor)
        self._revisar_combinada(celda)
        self._revisar_formula(celda)

        # El valor anterior es el que hay ahora: el del archivo, o el de una
        # escritura anterior a la misma celda dentro de esta misma sesión.
        if celda in self.cambios:
            anterior = self.cambios[celda]
        else:
            anterior = self._valores.get(celda)
        anterior = "" if anterior is None else anterior

        self.cambios[celda] = valor
        self.bitacora.append(
            {
                "hoja": HOJA_CAPTURA,
                "celda": celda,
                "valor_anterior": anterior,
                "valor_nuevo": valor,
                "documento": str(documento),
                "fecha_hora": datetime.now().isoformat(timespec="seconds"),
            }
        )
        return self.bitacora[-1]

    # -- guardado -----------------------------------------------------

    def guardar(self, recalcular_totales=True):
        """Aplica los cambios, verifica el resultado y escribe la bitácora.

        El orden importa: primero se escribe, después se verifica, y solo
        si la verificación pasa se deja el archivo y se escribe la bitácora.
        Si la verificación falla, el archivo se borra: es preferible
        quedarse sin archivo que entregar uno con las fórmulas rotas.

        Al final se le pide a LibreOffice que calcule los totales, para que
        el programa pueda leerlos y mostrarlos en pantalla. Con
        recalcular_totales=False se salta ese paso (sirve para las pruebas
        y para cuando solo interesa el archivo).

        Devuelve (ruta del archivo, ruta de la bitácora).
        """
        if self.guardado:
            raise EscrituraBloqueada("Este archivo ya se guardó.")

        _aplicar_cambios_al_zip(self.ruta_copia, self.cambios)

        # Regla 6. No hay forma de saltarse este paso ni de pedir que no se
        # haga: es parte de guardar.
        try:
            self.informe_verificacion = verificar_contra_original(
                self.ruta_original, self.ruta_copia
            )
        except VerificacionFallida:
            # El archivo quedó dudoso: se descarta. missing_ok porque si ya
            # no está, tampoco hay nada que borrar.
            self.ruta_copia.unlink(missing_ok=True)
            raise

        # Recálculo. Si LibreOffice no está o falla, no se rompe nada: se
        # entrega el archivo igual y el aviso explica que los totales se
        # ven al abrirlo en Excel.
        errores_previos = set()
        if recalcular_totales:
            errores_previos = set(_errores_recordados(self.ruta_original))
            self.informe_recalculo = recalcular(self.ruta_copia)

            if self.informe_recalculo["recalculado"]:
                # El archivo lo reescribió LibreOffice entero, así que se
                # vuelve a verificar. Si LibreOffice hubiera dañado alguna
                # fórmula, aquí se ve.
                try:
                    self.informe_verificacion = verificar_contra_original(
                        self.ruta_original, self.ruta_copia, tras_recalculo=True
                    )
                except VerificacionFallida:
                    self.ruta_copia.unlink(missing_ok=True)
                    raise

                # Errores dentro de las celdas (#REF!, #VALUE!...). Solo
                # cuentan los NUEVOS: si la plantilla ya traía uno, no es
                # culpa nuestra y no tiene sentido descartar por eso.
                nuevos = [
                    c for c in celdas_con_error(self.ruta_copia)
                    if c not in errores_previos
                ]
                if nuevos:
                    self.ruta_copia.unlink(missing_ok=True)
                    raise VerificacionFallida(
                        "Después de calcular los totales aparecieron"
                        f" {len(nuevos)} celdas con error. El archivo se"
                        " descartó y no se entrega.\n  - "
                        + "\n  - ".join(nuevos[:10])
                    )
                self.informe_verificacion["celdas_con_error"] = 0

        ruta_bitacora = self.ruta_copia.with_suffix(
            self.ruta_copia.suffix + SUFIJO_BITACORA
        )
        contenido = {
            "plantilla_original": str(self.ruta_original),
            "archivo_generado": str(self.ruta_copia),
            "hoja": HOJA_CAPTURA,
            "generado": datetime.now().isoformat(timespec="seconds"),
            "verificacion": self.informe_verificacion,
            "recalculo": self.informe_recalculo,
            "cambios": self.bitacora,
        }
        # ensure_ascii=False para que las tildes se vean; encoding explícito
        # para que Windows no lo lea en cp1252.
        ruta_bitacora.write_text(
            json.dumps(contenido, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )

        self.guardado = True
        return self.ruta_copia, ruta_bitacora


# ---------------------------------------------------------------------------
# La parte quirúrgica: tocar el XML de una sola hoja dentro del ZIP.
# ---------------------------------------------------------------------------


def _parte_de_la_hoja(zip_entrada, nombre_hoja):
    """Encuentra cuál archivo XML de adentro del ZIP es la hoja que se busca.

    El nombre de la hoja está en xl/workbook.xml, pero el archivo real
    (xl/worksheets/sheetN.xml) se averigua siguiendo la relación r:id. No se
    puede adivinar por el número: la quinta hoja no siempre es sheet5.xml.
    """
    workbook = zip_entrada.read("xl/workbook.xml").decode("utf-8")
    relaciones = zip_entrada.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    destinos = dict(
        re.findall(r'Id="(rId\d+)"[^>]*?Target="([^"]+)"', relaciones)
    )
    for nombre, rid in re.findall(
        r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', workbook
    ):
        if nombre == nombre_hoja:
            destino = destinos.get(rid, "")
            if not destino:
                break
            destino = destino.lstrip("/")
            if not destino.startswith("xl/"):
                destino = "xl/" + destino
            return destino

    raise EscrituraBloqueada(
        f"No se encontró la hoja «{nombre_hoja}» dentro del archivo de Excel."
    )


def _reemplazar_valor(xml, celda, valor):
    """Cambia el valor de UNA celda dentro del XML de la hoja.

    Las celdas vienen en tres formas y hay que manejar las tres:

        <c r="G32" s="499" t="n"><v>0</v></c>   celda con número
        <c r="G31" s="497"/>                     celda vacía
        <c r="G30" s="497" t="s"><v>12</v></c>   celda con texto

    De la celda se conserva el estilo (s="...") tal cual, que es el formato
    de número, los bordes y el color. Solo cambia el contenido.
    """
    patron = re.compile(
        r'<c r="%s"(?P<atributos>[^>]*?)(?:/>|>(?P<cuerpo>.*?)</c>)' % re.escape(celda),
        re.DOTALL,
    )
    encontrada = patron.search(xml)
    if encontrada is None:
        raise EscrituraBloqueada(
            f"La celda {celda} no existe en la hoja. No se inventa: se aborta."
        )

    cuerpo = encontrada.group("cuerpo") or ""
    # Segundo cerrojo de la regla 1, ahora sobre el XML crudo. El primero
    # fue con openpyxl. Si por lo que sea se cuela una fórmula hasta aquí,
    # se detiene igual.
    if "<f" in cuerpo:
        raise EscrituraBloqueada(
            f"{celda} contiene una fórmula en el archivo — escritura bloqueada."
        )

    # Conservar el estilo; quitar el tipo viejo y poner t="n" (número).
    atributos = encontrada.group("atributos")
    estilo = re.search(r'\ss="\d+"', atributos)
    nuevos_atributos = estilo.group(0) if estilo else ""

    reemplazo = '<c r="%s"%s t="n"><v>%s</v></c>' % (
        celda, nuevos_atributos, _formatear_numero(valor),
    )
    return xml[: encontrada.start()] + reemplazo + xml[encontrada.end():]


def _marcar_para_recalcular(workbook_xml):
    """Le pide a Excel que recalcule todo al abrir el archivo.

    Hace falta porque al cambiar una celda de captura, los resultados que
    quedaron guardados en las celdas de fórmula son de antes del cambio.
    Sin esta marca, el contador podría abrir el archivo y ver totales
    viejos. Es un ajuste de cálculo, no un cambio de contenido: no toca
    ninguna fórmula ni ningún dato.
    """
    if "fullCalcOnLoad" in workbook_xml:
        return workbook_xml
    if "<calcPr" in workbook_xml:
        return re.sub(
            r"<calcPr\b", '<calcPr fullCalcOnLoad="true"', workbook_xml, count=1
        )
    return workbook_xml.replace(
        "</workbook>", '<calcPr fullCalcOnLoad="true"/></workbook>'
    )


def _aplicar_cambios_al_zip(ruta_xlsx, cambios):
    """Reescribe el .xlsx copiando todo igual, menos las celdas pedidas.

    Se arma un archivo nuevo al lado y solo al final se reemplaza el
    anterior. Si algo falla a mitad de camino, el archivo de trabajo queda
    como estaba y no queda un .xlsx corrupto.
    """
    ruta_xlsx = Path(ruta_xlsx)
    ruta_temporal = ruta_xlsx.with_suffix(ruta_xlsx.suffix + ".enproceso")

    with zipfile.ZipFile(ruta_xlsx, "r") as entrada:
        parte_hoja = _parte_de_la_hoja(entrada, HOJA_CAPTURA)

        xml_hoja = entrada.read(parte_hoja).decode("utf-8")
        for celda, valor in cambios.items():
            xml_hoja = _reemplazar_valor(xml_hoja, celda, valor)

        xml_workbook = entrada.read("xl/workbook.xml").decode("utf-8")
        if cambios:
            xml_workbook = _marcar_para_recalcular(xml_workbook)

        nuevos = {
            parte_hoja: xml_hoja.encode("utf-8"),
            "xl/workbook.xml": xml_workbook.encode("utf-8"),
        }

        with zipfile.ZipFile(ruta_temporal, "w", zipfile.ZIP_DEFLATED) as salida:
            for info in entrada.infolist():
                if info.filename in nuevos:
                    # Se conserva la fecha original de la parte; lo único
                    # que cambia es el contenido.
                    nueva_info = zipfile.ZipInfo(
                        info.filename, date_time=info.date_time
                    )
                    nueva_info.compress_type = info.compress_type
                    nueva_info.external_attr = info.external_attr
                    salida.writestr(nueva_info, nuevos[info.filename])
                else:
                    # Todo lo demás pasa tal cual: imágenes, notas, estilos,
                    # las otras 14 hojas. No se abren ni se interpretan.
                    salida.writestr(info, entrada.read(info.filename))

    ruta_temporal.replace(ruta_xlsx)
    return ruta_xlsx


# ---------------------------------------------------------------------------
# Regla 6: la verificación posterior. Se corre siempre, después de guardar.
# ---------------------------------------------------------------------------


def leer_todas_las_formulas(ruta_xlsx):
    """Devuelve todas las fórmulas del libro: hoja -> {celda: fórmula}.

    Se abre con data_only=False, que es como se leen las fórmulas tal como
    están escritas. NUNCA con data_only=True: esa carga trae los resultados
    en vez de las fórmulas, y si se guardara, las borraría del archivo.
    """
    # read_only=True: openpyxl va leyendo el archivo por partes en vez de
    # armar el libro entero en memoria. Aquí solo se recorre y se lee, que
    # es justo para lo que sirve ese modo, y tarda 3,5 veces menos: 0,24
    # segundos en vez de 0,84. Se comprobó que devuelve exactamente las
    # mismas 902 fórmulas que la carga normal.
    libro = load_workbook(Path(ruta_xlsx), data_only=False, read_only=True)
    try:
        todas = {}
        for nombre in libro.sheetnames:
            hoja = libro[nombre]
            formulas = {}
            for fila in hoja.iter_rows():
                for celda in fila:
                    if es_formula(celda.value):
                        formulas[celda.coordinate] = celda.value
            todas[nombre] = formulas
        return todas
    finally:
        # Cerrar es obligatorio en este modo: deja el archivo abierto, y en
        # Windows un archivo abierto no se puede reemplazar. Sin esto,
        # LibreOffice no podría devolver el libro recalculado encima.
        libro.close()


# Leer las 902 fórmulas de un libro toma más de un segundo, y la plantilla
# original se lee en cada archivo que se genera. Se guarda lo leído en
# memoria, con la ruta, la fecha y el tamaño del archivo como llave: si
# alguien cambia la plantilla, la llave cambia y se vuelve a leer.
#
# El candado hace falta porque el servidor atiende cada petición en su
# propio hilo. Sin él, un hilo puede vaciar la memoria justo entre que otro
# guarda su resultado y lo lee de vuelta, y ese otro se cae con KeyError.
_recordado = {}
_candado_recordado = threading.Lock()
_CUANTOS_SE_RECUERDAN = 4


def _formulas_recordadas(ruta):
    """Como leer_todas_las_formulas, pero sin releer el mismo archivo dos veces."""
    ruta = Path(ruta)
    llave = llave_de_archivo(ruta)
    with _candado_recordado:
        if llave not in _recordado:
            if len(_recordado) >= _CUANTOS_SE_RECUERDAN:
                _recordado.clear()
            _recordado[llave] = leer_todas_las_formulas(ruta)
        return _recordado[llave]


def _errores_recordados(ruta):
    """Lo mismo, para las celdas con error de la plantilla original."""
    ruta = Path(ruta)
    llave = ("errores",) + llave_de_archivo(ruta)
    with _candado_recordado:
        if llave not in _recordado:
            if len(_recordado) >= _CUANTOS_SE_RECUERDAN:
                _recordado.clear()
            _recordado[llave] = celdas_con_error(ruta)
        return _recordado[llave]


def verificar_contra_original(ruta_original, ruta_generada,
                              tras_recalculo=False):
    """Compara el archivo generado con el original y devuelve un informe.

    Lanza VerificacionFallida si encuentra cualquier diferencia que no sea
    un valor de las celdas que se escribieron. Revisa cuatro cosas:

      1. Que estén las mismas hojas, con el mismo nombre y en el mismo orden.
      2. Que cada fórmula siga existiendo y diga exactamente lo mismo.
      3. Que no haya aparecido ninguna fórmula nueva.
      4. Que no se haya perdido ninguna parte interna del archivo
         (imágenes, notas, estilos). Esto último no lo pide la regla, pero
         es lo que se prometió al escribir de forma quirúrgica, así que se
         comprueba igual.

    tras_recalculo=True se usa cuando el archivo ya pasó por LibreOffice.
    En ese caso se salta una sola comprobación: la de que las partes
    internas estén byte por byte iguales. LibreOffice reescribe el archivo
    entero, así que ninguna parte queda idéntica aunque el contenido sea el
    mismo. Todo lo demás —las hojas, las 902 fórmulas y que no falte
    ninguna parte— se sigue exigiendo igual.
    """
    ruta_original = Path(ruta_original)
    ruta_generada = Path(ruta_generada)

    problemas = []

    formulas_originales = _formulas_recordadas(ruta_original)
    formulas_generadas = leer_todas_las_formulas(ruta_generada)

    hojas_originales = list(formulas_originales)
    hojas_generadas = list(formulas_generadas)
    if hojas_originales != hojas_generadas:
        problemas.append(
            f"Las hojas cambiaron. Antes: {hojas_originales}."
            f" Ahora: {hojas_generadas}."
        )

    total = 0
    diferencias = []
    for hoja in hojas_originales:
        antes = formulas_originales.get(hoja, {})
        ahora = formulas_generadas.get(hoja, {})
        total += len(antes)
        for celda, formula in antes.items():
            if celda not in ahora:
                diferencias.append(f"{hoja}!{celda}: desapareció la fórmula")
            elif ahora[celda] != formula:
                diferencias.append(
                    f"{hoja}!{celda}: decía «{formula}» y ahora dice"
                    f" «{ahora[celda]}»"
                )
        for celda in ahora:
            if celda not in antes:
                diferencias.append(
                    f"{hoja}!{celda}: apareció una fórmula que no estaba"
                )

    if diferencias:
        problemas.append(
            f"{len(diferencias)} fórmulas cambiaron. "
            + " | ".join(diferencias[:5])
        )

    # Partes internas del archivo.
    with zipfile.ZipFile(ruta_original) as original:
        partes_originales = set(original.namelist())
        try:
            parte_hoja = _parte_de_la_hoja(original, HOJA_CAPTURA)
        except EscrituraBloqueada:
            parte_hoja = ""
        contenido_original = {n: original.read(n) for n in partes_originales}

    with zipfile.ZipFile(ruta_generada) as generado:
        partes_generadas = set(generado.namelist())
        contenido_generado = {n: generado.read(n) for n in partes_generadas}

    perdidas = sorted(partes_originales - partes_generadas)
    if perdidas:
        problemas.append(
            f"Se perdieron {len(perdidas)} partes internas del archivo: "
            + ", ".join(perdidas[:5])
        )

    # Solo dos partes tienen permiso de cambiar: la hoja donde se escribe y
    # workbook.xml, por la marca de recalcular. Cualquier otra que cambie
    # es señal de que algo se reescribió sin querer.
    permitidas = {parte_hoja, "xl/workbook.xml"}
    intrusas = sorted(
        n for n in partes_originales & partes_generadas
        if n not in permitidas and contenido_original[n] != contenido_generado[n]
    )
    if intrusas and not tras_recalculo:
        problemas.append(
            f"Cambiaron {len(intrusas)} partes que no se debían tocar: "
            + ", ".join(intrusas[:5])
        )

    modificadas = sorted(
        n for n in partes_generadas
        if n not in contenido_original
        or contenido_original[n] != contenido_generado[n]
    )
    informe = {
        "formulas_comparadas": total,
        "formulas_distintas": len(diferencias),
        "hojas": len(hojas_originales),
        "partes_internas": len(partes_originales),
        "partes_modificadas_total": len(modificadas),
        # Solo las primeras: cuando el archivo pasa por LibreOffice quedan
        # modificadas casi todas, y la lista completa llenaría la bitácora
        # de ruido sin decir nada nuevo.
        "partes_modificadas": modificadas[:20],
        "problemas": problemas,
    }

    if problemas:
        raise VerificacionFallida(
            "El archivo generado NO pasó la verificación:\n  - "
            + "\n  - ".join(problemas)
        )

    return informe
