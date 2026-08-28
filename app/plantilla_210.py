"""
Mapa de la plantilla del Formulario 210.

Este módulo NO escribe nada. Solo abre la plantilla de Excel y devuelve un
mapa: para cada celda de las columnas de valores de la hoja de captura, dice
si es una celda donde se puede escribir o si tiene una fórmula que no se
puede tocar.

Por qué existe: la plantilla es un producto comercial con licencia de un
tercero. Las columnas de valores mezclan celdas de captura con celdas de
fórmula, sin ningún patrón fijo, así que hay que revisar celda por celda
antes de escribir. Este mapa es esa revisión.

La plantilla es un archivo del usuario, no un componente de este programa:
todo se hace por ruta, sin nombres de archivo quemados, para que funcione
igual con la plantilla que traiga cualquier otro contador.

Se corre solo, para ver el mapa:

    python -m app.plantilla_210 plantillas/mi_plantilla.xlsx
"""

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

RAIZ = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------------------
# Constantes de la plantilla.
#
# Están todas aquí arriba, en un solo lugar. Si otro contador trae una
# plantilla con otra hoja de captura, se cambia acá y no hay que salir a
# buscar el nombre regado por el resto del código.
# ---------------------------------------------------------------------------

# La ÚNICA hoja en la que este programa escribe. Las demás (el formulario
# oficial, los anexos, las tablas de impuestos y de reajuste fiscal) no se
# tocan nunca.
HOJA_CAPTURA = "Detalle renglón 210"

# Columnas de valores: subparciales, parciales y totales.
COLUMNAS_VALOR = ("G", "H", "I")

# Columnas donde puede estar la descripción del concepto. Se busca primero
# en F; si está vacía, en E. La plantilla cambia de layout a media hoja
# (hay un bloque donde el renglón está en D y la descripción en E), así que
# no sirve mirar una sola columna.
COLUMNAS_DESCRIPCION = ("F", "E")

# Columna de la sección (Patrimonio, Cédula General, Renta de pensiones...).
COLUMNA_SECCION = "C"

# En esa misma columna la plantilla también mete párrafos largos de
# explicación. Un nombre de sección es corto; si el texto pasa de este
# largo, no es una sección y la fila se queda con la sección anterior.
LARGO_MAXIMO_SECCION = 80

# Columnas donde puede aparecer el número de renglón del formulario.
COLUMNAS_RENGLON = ("D", "E")

# Columnas que se miran para saber hasta dónde llega el contenido de la hoja.
COLUMNAS_CONTENIDO = ("C", "D", "E", "F", "G", "H", "I", "J")

# Primera fila de datos. Arriba solo hay títulos, el presupuesto básico y
# los encabezados de la tabla ("Subparcial", "Parcial", "Totales").
FILA_INICIAL = 28

# Una fila cuya descripción empieza así es una nota al pie, no un concepto
# que se pueda diligenciar. Traen números adentro y engañan.
INICIOS_DE_NOTA = ("nota", "(nota")

# Los tres tipos de celda del mapa.
TIPO_FORMULA = "formula"
TIPO_CAPTURA = "captura"
TIPO_NO_APLICA = "no_aplica"


def es_formula(valor):
    """Dice si el contenido de una celda es una fórmula de Excel.

    openpyxl devuelve las fórmulas como texto que empieza por '='. Esta es
    la comprobación de la que depende todo lo demás, así que vive en una
    sola función y nadie la repite a mano.
    """
    return isinstance(valor, str) and valor.startswith("=")


def _texto(valor):
    """Convierte el contenido de una celda a texto limpio, sin saltos."""
    if valor is None:
        return ""
    return " ".join(str(valor).split())


def _mapa_de_combinadas(hoja):
    """Devuelve dos cosas sobre las celdas combinadas de la hoja.

    - 'ancla': para cada celda que forma parte de un grupo combinado, cuál
      es la celda de arriba a la izquierda (la única que guarda el valor).
    - En openpyxl, escribir en una celda combinada que no es el ancla lanza
      un error confuso. Además sería un error de todos modos: el valor no
      quedaría donde uno cree.
    """
    ancla = {}
    for rango in hoja.merged_cells.ranges:
        col_min, fila_min, col_max, fila_max = range_boundaries(str(rango))
        coord_ancla = f"{get_column_letter(col_min)}{fila_min}"
        for fila in range(fila_min, fila_max + 1):
            for col in range(col_min, col_max + 1):
                ancla[f"{get_column_letter(col)}{fila}"] = coord_ancla
    return ancla


def _valor_efectivo(hoja, coordenada, anclas):
    """El valor de una celda, resolviendo las combinadas.

    Si la celda es parte de un grupo combinado, el valor real está en la
    esquina de arriba a la izquierda. Sin esto, la sección y el número de
    renglón salen vacíos en casi todas las filas.
    """
    return hoja[anclas.get(coordenada, coordenada)].value


def _ultima_fila_con_contenido(hoja):
    """Hasta qué fila llega la tabla de verdad.

    No se quema un número fijo: Excel suele reportar miles de filas de más
    (formatos sobrantes), y otra plantilla tendrá otro largo. Se busca la
    última fila que tenga algo en alguna de las columnas de contenido.
    """
    ultima = FILA_INICIAL
    for fila in range(FILA_INICIAL, hoja.max_row + 1):
        for col in COLUMNAS_CONTENIDO:
            if hoja[f"{col}{fila}"].value not in (None, ""):
                ultima = fila
                break
    return ultima


def _numero_de_renglon(texto):
    """Si el texto es un número de renglón del formulario, lo devuelve.

    Los renglones del 210 van del 29 al 84 más algunos informativos. Se
    aceptan solo textos que sean puramente un número; cualquier otra cosa
    (una descripción, un nombre de subcédula) se descarta.
    """
    limpio = texto.strip()
    if limpio.isdigit():
        return limpio
    return ""


def mapear_plantilla(ruta_plantilla):
    """Lee la plantilla y devuelve el mapa de celdas de la hoja de captura.

    NO escribe, NO guarda y NO copia el archivo. Abre, lee y cierra.

    El libro se abre dos veces a propósito:

    1. Con las fórmulas tal como están escritas. Esta es la carga que manda:
       es la que dice si una celda tiene fórmula o no.
    2. Con data_only=True, que trae el último resultado que Excel calculó y
       guardó. Sirve para mostrar cuánto vale hoy cada total y para leer los
       números de renglón que salen de una fórmula. Esta segunda carga
       NUNCA se guarda: si se guardara, reemplazaría las 902 fórmulas del
       libro por números y las destruiría para siempre.

    Devuelve un diccionario con los datos de la hoja y la lista de celdas.
    """
    ruta = Path(ruta_plantilla)
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró la plantilla: {ruta}")

    libro = load_workbook(ruta, data_only=False)
    if HOJA_CAPTURA not in libro.sheetnames:
        raise ValueError(
            f"La plantilla no tiene la hoja «{HOJA_CAPTURA}». "
            f"Hojas encontradas: {', '.join(libro.sheetnames)}"
        )
    hoja = libro[HOJA_CAPTURA]

    # Segunda carga, solo de lectura, para los valores ya calculados.
    libro_calculado = load_workbook(ruta, data_only=True)
    hoja_calculada = libro_calculado[HOJA_CAPTURA]

    anclas = _mapa_de_combinadas(hoja)
    fila_final = _ultima_fila_con_contenido(hoja)

    celdas = []
    seccion_actual = ""

    for fila in range(FILA_INICIAL, fila_final + 1):
        # Sección: se arrastra hacia abajo. La plantilla la escribe una sola
        # vez, en la primera fila del bloque, y el resto queda en blanco.
        seccion = _texto(_valor_efectivo(hoja, f"{COLUMNA_SECCION}{fila}", anclas))
        if seccion and len(seccion) <= LARGO_MAXIMO_SECCION:
            seccion_actual = seccion

        # Descripción: primero F, y si está vacía, E.
        descripcion = ""
        columna_descripcion = ""
        for col in COLUMNAS_DESCRIPCION:
            coord = f"{col}{fila}"
            texto = _texto(_valor_efectivo(hoja, coord, anclas))
            # Un número suelto en E es el renglón, no una descripción.
            if texto and not _numero_de_renglon(texto):
                descripcion = texto
                # Si F está combinada con E, el texto vive en E: se anota la
                # columna donde de verdad está, no donde se encontró.
                columna_descripcion = anclas.get(coord, coord)[0]
                break

        # Número de renglón: puede estar en D o en E, y a veces es una
        # fórmula (por ejemplo "=+E103+1"), así que se usa el valor
        # calculado cuando el escrito no sirve.
        renglon = ""
        for col in COLUMNAS_RENGLON:
            if col == columna_descripcion:
                continue
            crudo = _texto(_valor_efectivo(hoja, f"{col}{fila}", anclas))
            if es_formula(crudo):
                crudo = _texto(hoja_calculada[f"{col}{fila}"].value)
            renglon = _numero_de_renglon(crudo)
            if renglon:
                break

        es_nota = descripcion.lower().startswith(INICIOS_DE_NOTA)

        for col in COLUMNAS_VALOR:
            coordenada = f"{col}{fila}"
            celda = hoja[coordenada]
            valor = celda.value
            formula = valor if es_formula(valor) else ""

            ancla = anclas.get(coordenada)
            combinada = ancla is not None
            es_ancla = (not combinada) or ancla == coordenada

            # Clasificación. El orden importa: la fórmula manda sobre todo
            # lo demás, porque es la regla que protege el archivo.
            if formula:
                tipo = TIPO_FORMULA
                motivo = "tiene fórmula"
            elif combinada and not es_ancla:
                tipo = TIPO_NO_APLICA
                motivo = f"celda combinada, el valor va en {ancla}"
            elif es_nota:
                tipo = TIPO_NO_APLICA
                motivo = "fila de nota al pie"
            elif not descripcion:
                tipo = TIPO_NO_APLICA
                motivo = "fila sin descripción (separador en blanco)"
            else:
                tipo = TIPO_CAPTURA
                motivo = ""

            celdas.append(
                {
                    "hoja": HOJA_CAPTURA,
                    "celda": coordenada,
                    "fila": fila,
                    "columna": col,
                    "tipo": tipo,
                    "motivo": motivo,
                    "valor_actual": "" if valor is None else valor,
                    # La plantilla trae un 0 escrito en las casillas que de
                    # verdad espera que alguien diligencie. Es la pista más
                    # confiable de "aquí va un dato" y la convención del
                    # propio archivo para limpiar un valor.
                    "cero_precargado": valor == 0,
                    "valor_calculado": hoja_calculada[coordenada].value,
                    "formula": formula,
                    "descripcion": descripcion,
                    "columna_descripcion": columna_descripcion,
                    "renglon": renglon,
                    "seccion": seccion_actual,
                    "combinada": combinada,
                }
            )

    return {
        "ruta": str(ruta),
        "hoja": HOJA_CAPTURA,
        "hojas_del_libro": list(libro.sheetnames),
        "fila_inicial": FILA_INICIAL,
        "fila_final": fila_final,
        "celdas": celdas,
    }


def contar_formulas_del_libro(ruta_plantilla):
    """Cuenta las fórmulas de TODO el libro, hoja por hoja.

    Es la foto de referencia: en el Paso 3, después de escribir y guardar,
    se vuelve a contar y se compara. Si cambió una sola fórmula, el
    resultado se descarta.
    """
    libro = load_workbook(Path(ruta_plantilla), data_only=False)
    conteo = {}
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        total = 0
        for fila in hoja.iter_rows():
            for celda in fila:
                if es_formula(celda.value):
                    total += 1
        conteo[nombre] = total
    return conteo


def resumen(mapa):
    """Cuenta cuántas celdas hay de cada tipo."""
    conteo = {TIPO_CAPTURA: 0, TIPO_FORMULA: 0, TIPO_NO_APLICA: 0}
    for celda in mapa["celdas"]:
        conteo[celda["tipo"]] += 1
    return conteo


def guardar_csv(mapa, ruta_csv):
    """Escribe el mapa a un CSV para poder revisarlo con calma en Excel.

    utf-8-sig (no utf-8 a secas) porque es lo que hace que Excel en Windows
    muestre bien las tildes al abrir el archivo con doble clic.
    """
    ruta = Path(ruta_csv)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    columnas = [
        "celda", "tipo", "motivo", "seccion", "renglon", "descripcion",
        "columna_descripcion", "valor_actual", "valor_calculado", "formula",
        "cero_precargado", "combinada",
    ]
    with open(ruta, "w", encoding="utf-8-sig", newline="") as archivo:
        escritor = csv.DictWriter(
            archivo, fieldnames=columnas, extrasaction="ignore"
        )
        escritor.writeheader()
        for celda in mapa["celdas"]:
            escritor.writerow(celda)
    return ruta


def imprimir_mapa(mapa, tipos=(TIPO_CAPTURA, TIPO_FORMULA), limite=40):
    """Muestra el mapa en la terminal, en líneas cortas.

    Por defecto muestra solo las celdas que importan (captura y fórmula) y
    corta a 40 por tipo, para que no se pierda al desplazarse. Con
    limite=None muestra todas.
    """
    conteo = resumen(mapa)
    print(f"Plantilla : {mapa['ruta']}")
    print(f"Hoja      : {mapa['hoja']}  (de {len(mapa['hojas_del_libro'])} hojas)")
    print(f"Filas     : {mapa['fila_inicial']} a {mapa['fila_final']}")
    print(f"Celdas    : {len(mapa['celdas'])} en columnas {', '.join(COLUMNAS_VALOR)}")
    print()
    con_cero = sum(
        1 for c in mapa["celdas"]
        if c["tipo"] == TIPO_CAPTURA and c["cero_precargado"]
    )
    print(f"  de captura : {conteo[TIPO_CAPTURA]}   ({con_cero} traen un 0 escrito)")
    print(f"  con fórmula: {conteo[TIPO_FORMULA]}   (bloqueadas)")
    print(f"  no aplica  : {conteo[TIPO_NO_APLICA]}")

    for tipo in tipos:
        del_tipo = [c for c in mapa["celdas"] if c["tipo"] == tipo]
        mostradas = del_tipo if limite is None else del_tipo[:limite]
        print()
        print(f"--- {tipo.upper()} ({len(del_tipo)}) ---")
        for celda in mostradas:
            renglon = f"r{celda['renglon']}" if celda["renglon"] else "  -"
            detalle = celda["formula"] or celda["descripcion"]
            print(f"{celda['celda']:>6} {renglon:>4}  {detalle[:58]}")
        if limite is not None and len(del_tipo) > limite:
            print(f"       ... y {len(del_tipo) - limite} más (están en el CSV)")


def _principal():
    """Lo que pasa al correr 'python -m app.plantilla_210'."""
    if len(sys.argv) > 1:
        ruta = Path(sys.argv[1])
    else:
        # Sin argumento: la única plantilla que haya en plantillas/.
        encontradas = sorted(
            p for p in (RAIZ / "plantillas").glob("*.xlsx")
            if not p.name.startswith("~$")
        )
        if not encontradas:
            print("No hay ninguna plantilla .xlsx en la carpeta plantillas/")
            return 1
        ruta = encontradas[0]

    mapa = mapear_plantilla(ruta)
    imprimir_mapa(mapa)

    ruta_csv = guardar_csv(mapa, RAIZ / "datos" / "mapa_plantilla_210.csv")
    print()
    print(f"Mapa completo guardado en: {ruta_csv.relative_to(RAIZ)}")
    print("La plantilla no se modificó: este paso solo lee.")
    return 0


if __name__ == "__main__":
    sys.exit(_principal())
