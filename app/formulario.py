"""
El Formulario 210 de cada cliente.

La plantilla es una sola y es de todos: el archivo con licencia que el
contador dejó en la carpeta plantillas/. Lo que cambia de un cliente a
otro son los valores que se capturan.

Cómo se guarda cada cliente por separado
----------------------------------------
Lo que se guarda en la base de datos NO es el archivo de Excel, son los
valores: "para el cliente 3, en la celda G32 va 1.500.000, y ese dato
salió de tal documento". El archivo de Excel se arma cuando alguien lo
pide, así:

    plantilla limpia  ->  copia nueva  ->  se escriben los valores
    del cliente       ->  se verifica  ->  se calculan los totales

Esto tiene tres ventajas sobre ir editando un archivo por cliente:

  1. Cada archivo sale siempre de la plantilla original, sin tocar. Nunca
     se escribe encima de un archivo que ya se escribió antes, así que no
     se van acumulando reescrituras.
  2. La verificación completa se puede hacer siempre, porque siempre hay
     con qué comparar: la plantilla limpia.
  3. Corregir o quitar un valor es cambiar una fila de la base. El archivo
     se vuelve a armar y ya.

Cada cliente tiene su carpeta en datos/formularios/cliente-N/, y ahí queda
su archivo con su bitácora. El de un cliente nunca se mezcla con el de otro.
"""

import re
import shutil
import threading
import unicodedata
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from app import db
from app.documentos import sanitizar_nombre
from app.escribir_210 import EscritorPlantilla, EscrituraBloqueada
from app.plantilla_210 import (
    HOJA_CAPTURA, RAIZ, TIPO_CAPTURA, mapear_plantilla,
)
from app.recalcular import buscar_libreoffice, valores_calculados_del_libro

CARPETA_PLANTILLAS = RAIZ / "plantillas"
CARPETA_FORMULARIOS = RAIZ / "datos" / "formularios"

# El archivo de cada cliente siempre se llama igual dentro de su carpeta.
# El nombre bonito (con el nombre del cliente) se le pone al descargarlo.
NOMBRE_ARCHIVO = "formulario.xlsx"

# Cuántos resultados devuelve una búsqueda de conceptos.
LIMITE_BUSQUEDA = 40

# Cuántos conceptos padre se muestran encima de una casilla, y qué tan
# largo puede ser cada uno. La plantilla tiene títulos larguísimos; en
# pantalla lo que sirve es el rastro corto: "Gastos personales › Ropa".
CUANTOS_PADRES = 2
LARGO_DEL_PADRE = 38

# Secciones cuyos totales NO se muestran en pantalla.
#
# "Liquidación Privada" son los renglones 116 a 137: el impuesto a cargo,
# el anticipo y los saldos a pagar o a favor. Este programa no calcula
# impuestos ni dice cuánto hay que pagar, así que tampoco los muestra en
# pantalla. Están en el archivo de Excel, que es la herramienta del
# contador, y ahí los ve él.
SECCIONES_QUE_NO_SE_MUESTRAN = ("Liquidación Privada",)

# En qué ajuste se guarda cuál plantilla está en uso.
CLAVE_PLANTILLA = "plantilla_activa"

# Tope de tamaño para una plantilla subida. Las de Excel con anexos pesan
# uno o dos megas; veinte es de sobra y evita que alguien suba un archivo
# gigante por error.
MAXIMO_PLANTILLA = 20 * 1024 * 1024


class SinPlantilla(Exception):
    """No hay ninguna plantilla en la carpeta plantillas/."""


# ---------------------------------------------------------------------------
# La plantilla y su mapa
# ---------------------------------------------------------------------------


def listar_plantillas():
    """Todas las plantillas que hay en la carpeta plantillas/.

    Se ignoran los archivos que empiezan por ~$: son los temporales que
    Excel deja cuando alguien tiene el archivo abierto.
    """
    if not CARPETA_PLANTILLAS.exists():
        return []
    return sorted(
        p for p in CARPETA_PLANTILLAS.glob("*.xlsx")
        if not p.name.startswith("~$")
    )


def ruta_plantilla():
    """La plantilla que está en uso, o None si no hay ninguna.

    El contador puede tener varias y elegir cuál usa. Si la que eligió ya
    no está (la borró, la renombró), se usa la primera que haya en vez de
    quedarse sin funcionar.
    """
    encontradas = listar_plantillas()
    if not encontradas:
        return None

    elegida = db.leer_ajuste(CLAVE_PLANTILLA)
    for candidata in encontradas:
        if candidata.name == elegida:
            return candidata
    return encontradas[0]


def elegir_plantilla(nombre):
    """Deja marcada cuál plantilla se usa de ahora en adelante."""
    nombre = (nombre or "").strip()
    for candidata in listar_plantillas():
        if candidata.name == nombre:
            db.guardar_ajuste(CLAVE_PLANTILLA, nombre)
            return candidata
    raise SinPlantilla(f"No hay ninguna plantilla que se llame «{nombre}».")


def revisar_plantilla(contenido):
    """Revisa que un archivo subido sirva como plantilla. Si no, explica por qué.

    Se pide una sola cosa: que tenga la hoja de captura. Todo lo demás
    (cuántas fórmulas, cuántos anexos) puede cambiar de una plantilla a
    otra y no es asunto nuestro.
    """
    if len(contenido) > MAXIMO_PLANTILLA:
        raise SinPlantilla(
            "El archivo pesa más de 20 MB. ¿Seguro que es la plantilla?"
        )
    if not zipfile.is_zipfile(BytesIO(contenido)):
        raise SinPlantilla(
            "Ese archivo no es un Excel (.xlsx). Si es un .xls viejo,"
            " ábralo en Excel y guárdelo como .xlsx."
        )

    try:
        libro = load_workbook(BytesIO(contenido), data_only=False)
    except Exception:
        raise SinPlantilla(
            "No se pudo abrir el archivo. Puede estar dañado o protegido"
            " con contraseña."
        )

    if HOJA_CAPTURA not in libro.sheetnames:
        raise SinPlantilla(
            f"La plantilla no tiene la hoja «{HOJA_CAPTURA}», que es donde"
            f" el programa escribe. Hojas que trae:"
            f" {', '.join(libro.sheetnames[:8])}."
        )
    return libro.sheetnames


def guardar_plantilla_subida(nombre_original, contenido):
    """Guarda una plantilla que subió el contador y la deja en uso.

    El archivo se guarda tal cual llegó: es de él, con su licencia. No se
    abre para modificarlo ni se le quita nada.
    """
    revisar_plantilla(contenido)

    nombre = sanitizar_nombre(nombre_original or "plantilla.xlsx")
    if not nombre.lower().endswith(".xlsx"):
        nombre += ".xlsx"

    CARPETA_PLANTILLAS.mkdir(parents=True, exist_ok=True)
    destino = CARPETA_PLANTILLAS / nombre

    # Si ya hay una con ese nombre, se le agrega la fecha en vez de pisarla.
    if destino.exists():
        marca = datetime.now().strftime("%Y%m%d-%H%M")
        destino = CARPETA_PLANTILLAS / f"{Path(nombre).stem} ({marca}).xlsx"

    destino.write_bytes(contenido)
    db.guardar_ajuste(CLAVE_PLANTILLA, destino.name)
    return destino


# Leer la plantilla y armar el mapa toma un segundo y medio. Se guarda en
# memoria para no repetirlo en cada tecla que el contador escribe en el
# buscador. Si cambia el archivo (otra plantilla, o la misma corregida), la
# fecha de modificación cambia y el mapa se vuelve a armar solo.
#
# El candado hace falta porque el servidor atiende cada petición en su
# propio hilo. Sin él, dos hilos pueden estar armando el mapa a la vez y
# uno puede leer el diccionario cuando el otro lo dejó a medio actualizar.
_mapa_guardado = {"ruta": None, "modificado": None, "mapa": None, "indice": None}

# Y lo mismo para las celdas que alguna fórmula lee: recorrer todas las
# fórmulas de la plantilla se demora, y no cambia hasta que cambie el
# archivo.
_leidas_guardadas = {"ruta": None, "modificado": None, "celdas": set()}
_candado_leidas = threading.Lock()
_candado_mapa = threading.Lock()


def mapa():
    """El mapa de celdas de la plantilla, sacado de memoria si ya se leyó."""
    return _mapa_e_indice()[0]


def indice():
    """Las celdas del mapa en un diccionario: {'G32': {...}}."""
    return _mapa_e_indice()[1]


def _mapa_e_indice():
    """El mapa y su índice, armados una sola vez aunque los pidan a la vez."""
    ruta = ruta_plantilla()
    if ruta is None:
        raise SinPlantilla(
            "No hay ninguna plantilla en la carpeta plantillas/."
            " Ponga ahí el archivo de Excel del Formulario 210."
        )

    # st_mtime_ns y no st_mtime: en Windows la fecha en segundos puede
    # repetirse entre dos escrituras seguidas, y entonces se seguiría
    # usando el mapa de la plantilla anterior.
    modificado = ruta.stat().st_mtime_ns
    with _candado_mapa:
        if (_mapa_guardado["ruta"] != ruta
                or _mapa_guardado["modificado"] != modificado):
            armado = mapear_plantilla(ruta)
            _mapa_guardado.update(
                ruta=ruta,
                modificado=modificado,
                mapa=armado,
                indice={c["celda"]: c for c in armado["celdas"]},
            )
        return _mapa_guardado["mapa"], _mapa_guardado["indice"]


# Para encontrar las celdas que menciona una fórmula: G115, $H$114,
# G115:G120, 'Otra hoja'!G12.
#
# El nombre de la hoja importa y no es un detalle: la plantilla tiene
# varias hojas y en otra hay un =SUM(G113:G124) que no tiene nada que
# ver con la hoja de captura. Sin distinguirlas, esa suma hacía parecer
# «conectada» una casilla que no lo está.
_REFERENCIA = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_.\u00C0-\u024F]+))!)?"
    r"\$?([A-Z]{1,3})\$?(\d{1,7})"
    r"(?::\$?([A-Z]{1,3})\$?(\d{1,7}))?")

# Un rango más grande que esto es una barrida de toda la hoja, no una
# suma de renglones: no se expande, gastaría memoria sin aportar nada.
_RANGO_MAXIMO = 20000


def celdas_que_alguien_lee():
    """Las casillas que alguna fórmula de la plantilla SÍ lee.

    Esto no es un detalle técnico: es la diferencia entre anotar una
    cifra y creer que se anotó.

    La plantilla del contador ofrece tres columnas de valor en cada
    fila —Subparcial, Parcial y Totales— pero en cada renglón solo UNA
    está cableada a las sumas. El renglón 32, por ejemplo, saca su
    total de `H114+H123`, que a su vez suman la columna G de las filas
    de detalle. Las casillas G113 y H113, que están justo al lado del
    título «Ingresos brutos por rentas de trabajo», no las lee ninguna
    fórmula: escribir ahí deja el número quieto y el renglón en cero.

    De 1.317 casillas donde se puede escribir, solo unas 318 están
    conectadas. Ofrecer las otras es ofrecer un error silencioso.
    """
    ruta = ruta_plantilla()
    if ruta is None:
        raise SinPlantilla(
            "No hay ninguna plantilla en la carpeta plantillas/."
        )

    modificado = ruta.stat().st_mtime_ns
    with _candado_leidas:
        if (_leidas_guardadas["ruta"] == ruta
                and _leidas_guardadas["modificado"] == modificado):
            return _leidas_guardadas["celdas"]

        libro = load_workbook(ruta, data_only=False)
        leidas = set()
        for hoja in libro.worksheets:
            for fila in hoja.iter_rows():
                for celda in fila:
                    valor = celda.value
                    if not (isinstance(valor, str) and valor.startswith("=")):
                        continue
                    for encontrada in _REFERENCIA.finditer(valor):
                        _sumar_referencia(leidas, encontrada, hoja.title)
        libro.close()

        _leidas_guardadas.update(ruta=ruta, modificado=modificado,
                                 celdas=leidas)
        return leidas


def _sumar_referencia(leidas, encontrada, hoja_de_la_formula):
    """Agrega una celda suelta, o todas las de un rango.

    Solo cuentan las de la hoja de captura. Una referencia sin nombre de
    hoja es de la hoja donde está la fórmula.
    """
    entrecomillada, suelta, col1, fila1, col2, fila2 = encontrada.groups()
    nombre_hoja = entrecomillada or suelta or hoja_de_la_formula
    if nombre_hoja != HOJA_CAPTURA:
        return
    if col2 is None:
        leidas.add(col1 + fila1)
        return
    try:
        min_col, min_fila, max_col, max_fila = range_boundaries(
            "%s%s:%s%s" % (col1, fila1, col2, fila2))
    except Exception:
        return
    if (max_col - min_col + 1) * (max_fila - min_fila + 1) > _RANGO_MAXIMO:
        return
    for columna in range(min_col, max_col + 1):
        for fila in range(min_fila, max_fila + 1):
            leidas.add(get_column_letter(columna) + str(fila))


def _clave_de_recordada(numero):
    """Con qué llave se recuerda la casilla que el contador escogió.

    Lleva el nombre de la plantilla adentro: si él cambia de plantilla,
    las casillas de la anterior no significan lo mismo y no se le pueden
    volver a proponer.
    """
    ruta = ruta_plantilla()
    return "casilla_renglon_%s_%s" % (ruta.name if ruta else "", numero)


def recordar_casilla(numero, celda):
    """Se acuerda de en qué casilla puso él este renglón la vez pasada.

    Es lo mismo que hace el programa con las correcciones de
    clasificación: la decisión es suya, y una vez tomada no hay por qué
    volver a preguntársela cada vez.
    """
    if numero and celda:
        db.guardar_ajuste(_clave_de_recordada(numero), celda)


def casilla_recordada(numero):
    """La casilla que él escogió antes para este renglón, o ''."""
    if not numero:
        return ""
    return db.leer_ajuste(_clave_de_recordada(numero), "")


def elegir_casilla(numero, texto=""):
    """Cuál casilla del renglón le corresponde a este dato.

    Devuelve (recomendada, motivo, todas). `recomendada` puede ser None,
    y entonces le toca escoger al contador.

    Esto NO es una decisión tributaria y por eso el programa sí la
    puede tomar: el renglón ya está decidido: lo decidió la DIAN en la
    exógena o lo decidió él. Lo que falta es en cuál fila de SU hoja de
    trabajo va, y eso se resuelve leyendo la etiqueta de la fila:
    «Pagos por salarios» contra una fila que dice «Salarios».

    Tres caminos, en este orden:
      1. La que él ya escogió antes para este renglón.
      2. La única que hay, si solo hay una.
      3. La fila cuya descripción se parece al concepto del documento.
    """
    todas = celdas_de_renglon(numero)
    if not todas:
        return None, "", []

    recordada = casilla_recordada(numero)
    for celda in todas:
        if celda["celda"] == recordada:
            return celda, "Es la que usted usó la vez pasada.", todas

    if len(todas) == 1:
        return todas[0], "Es la única casilla de ese renglón.", todas

    if texto:
        parecida = _la_mas_parecida(texto, todas)
        if parecida is not None:
            return parecida, ("La fila de la plantilla dice «%s»."
                              % (parecida["descripcion"] or "")[:60]), todas

    return None, "", todas


def _la_mas_parecida(texto, casillas):
    """La casilla cuya descripción se parece más al texto. None si empatan."""
    from app import lectura

    del_dato = lectura.palabras_utiles(texto)
    if not del_dato:
        return None

    puntajes = []
    for casilla in casillas:
        comunes = del_dato & lectura.palabras_utiles(casilla["descripcion"])
        if comunes:
            puntajes.append((len(comunes), casilla))
    if not puntajes:
        return None
    puntajes.sort(key=lambda par: par[0], reverse=True)
    # Empate: no hay forma de saber cuál, y adivinar aquí es escribir una
    # cifra en el sitio equivocado de su declaración.
    if len(puntajes) > 1 and puntajes[1][0] == puntajes[0][0]:
        return None
    return puntajes[0][1]


def celdas_de_renglon(numero):
    """Las casillas donde de verdad se escribe un renglón del 210.

    'numero' es el del formulario sin la R: "32" para R32.

    Dos cosas que antes no hacía, y las dos eran errores:

    1. Busca en TODO el bloque del renglón, no solo en la fila que
       lleva su número. En la plantilla del contador esa fila es un
       título y la cifra se escribe en las filas de detalle de abajo:
       «Salarios», «Cesantías e intereses», «Empresa 1, NIT…».

    2. Deja fuera las casillas que ninguna fórmula lee. Escribir en una
       de esas no rompe nada, y eso es lo peligroso: parece que quedó
       anotado y el renglón se queda en cero.

    Devuelve las casillas en el orden en que están en la hoja, que es
    el orden en que el contador las lee.
    """
    numero = str(numero).strip()
    if not numero:
        return []

    celdas = mapa()["celdas"]
    leidas = celdas_que_alguien_lee()

    # El bloque va desde la fila del renglón hasta la del SIGUIENTE
    # renglón distinto.
    #
    # Se mira «distinto» y no «cualquiera» porque la plantilla es
    # inconsistente: en unos bloques la casilla del número está
    # combinada y todas sus filas repiten el mismo renglón (el 29 lo
    # repite en sesenta filas), y en otros el número aparece una sola
    # vez y las filas de abajo van en blanco (el 32). Con esta regla los
    # dos casos salen bien.
    fila_inicio = None
    fila_fin = None
    for celda in celdas:
        suyo = celda.get("renglon") or ""
        if fila_inicio is None:
            if suyo == numero:
                fila_inicio = celda["fila"]
            continue
        if celda["fila"] <= fila_inicio:
            continue
        if suyo and suyo != numero:
            fila_fin = celda["fila"]
            break
    if fila_inicio is None:
        return []

    encontradas = []
    for celda in celdas:
        if celda["tipo"] != TIPO_CAPTURA:
            continue
        if celda["fila"] < fila_inicio:
            continue
        if fila_fin is not None and celda["fila"] >= fila_fin:
            continue
        if celda["celda"] not in leidas:
            # Nadie la lee: ofrecerla sería ofrecer un error silencioso.
            continue
        encontradas.append(_para_pantalla(celda))
    return encontradas


def resumen_plantilla():
    """Lo que la pantalla necesita saber sobre la plantilla."""
    ruta = ruta_plantilla()
    if ruta is None:
        return {
            "hay_plantilla": False,
            "motivo": (
                "Todavía no hay ninguna plantilla. Suba su archivo de Excel"
                " del Formulario 210 con el botón de abajo, o déjelo en la"
                " carpeta plantillas/ del programa."
            ),
            "libreoffice": bool(buscar_libreoffice()),
        }

    armado = mapa()
    de_captura = sum(
        1 for c in armado["celdas"] if c["tipo"] == TIPO_CAPTURA
    )
    return {
        "hay_plantilla": True,
        "archivo": ruta.name,
        "hoja": armado["hoja"],
        "celdas_de_captura": de_captura,
        "libreoffice": bool(buscar_libreoffice()),
        "disponibles": [p.name for p in listar_plantillas()],
    }


def buscar_celdas(texto, limite=LIMITE_BUSQUEDA, solo_esperadas=True):
    """Busca conceptos de la plantilla por palabra o por número de renglón.

    Solo devuelve celdas donde se puede escribir.

    Por defecto muestra únicamente las casillas que la plantilla trae con
    un 0 puesto, que son las que el archivo espera que alguien diligencie.
    En una fila hay tres columnas escribibles (subparcial, parcial y total)
    y casi siempre el dato va en una sola; mostrar las tres confunde más de
    lo que ayuda. Con solo_esperadas=False se ven todas, para el caso raro
    en que el contador necesite otra.
    """
    texto = _sin_tildes(texto).strip()
    encontradas = []

    for celda in mapa()["celdas"]:
        if celda["tipo"] != TIPO_CAPTURA:
            continue
        if solo_esperadas and not celda["cero_precargado"]:
            continue
        if texto:
            en_descripcion = texto in _sin_tildes(celda["descripcion"])
            en_contexto = any(
                texto in _sin_tildes(c) for c in (celda.get("contexto") or [])
            )
            en_renglon = texto == celda["renglon"]
            en_celda = texto == celda["celda"].lower()
            en_seccion = texto in _sin_tildes(celda["seccion"])
            if not (en_descripcion or en_contexto or en_renglon or en_celda
                    or en_seccion):
                continue
        encontradas.append(celda)

    encontradas.sort(
        key=lambda c: (not c["cero_precargado"], c["fila"], c["columna"])
    )
    return [_para_pantalla(c) for c in encontradas[:limite]]


def _sin_tildes(texto):
    """'Alimentación' -> 'alimentacion', para poder buscar sin tildes.

    El contador escribe rápido y no siempre pone las tildes. Que la
    búsqueda no le sirva por eso sería absurdo.
    """
    sin_marcas = unicodedata.normalize("NFD", (texto or "").lower())
    return "".join(c for c in sin_marcas if unicodedata.category(c) != "Mn")


def _rastro(contexto):
    """Los conceptos padre, cortos y en una sola línea.

    Los títulos de la plantilla son larguísimos y casi siempre la parte
    útil está antes del primer paréntesis o de la primera coma:

        "Gastos personales (incluido el IVA o INC que…)" -> "Gastos personales"

    Si cortar por ahí deja algo demasiado corto, se recorta a lo bruto.
    """
    cortos = []
    for texto in (contexto or [])[-CUANTOS_PADRES:]:
        for signo in ("(", ";", ":", ","):
            if signo in texto:
                antes = texto.split(signo)[0].strip()
                if len(antes) >= 12:
                    texto = antes
                    break
        if len(texto) > LARGO_DEL_PADRE:
            texto = texto[:LARGO_DEL_PADRE].rstrip(" ,;:(") + "…"
        cortos.append(texto)
    return " › ".join(cortos)


def _para_pantalla(celda):
    """Deja de una celda del mapa solo lo que la pantalla usa."""
    return {
        "celda": celda["celda"],
        "columna": celda["columna"],
        "fila": celda["fila"],
        "descripcion": celda["descripcion"],
        # De qué conceptos cuelga: "Gastos personales › Alimentación".
        "contexto": _rastro(celda.get("contexto")),
        "seccion": celda["seccion"],
        "renglon": celda["renglon"],
        "esperado": celda["cero_precargado"],
    }


# ---------------------------------------------------------------------------
# Los valores de cada cliente
# ---------------------------------------------------------------------------


def valores_calculados(ruta_xlsx):
    """Todos los valores ya calculados de la hoja de captura: {celda: valor}.

    La lectura del libro la hace `valores_calculados_del_libro`, que la
    guarda en memoria: el mismo archivo no se vuelve a abrir aunque se le
    pregunten cosas distintas. Aquí solo se escoge la hoja de captura.
    """
    return valores_calculados_del_libro(ruta_xlsx).get(HOJA_CAPTURA, {})


def _numero_o_nada(valor):
    """Devuelve el número, o None si eso no es un número."""
    if isinstance(valor, bool) or not isinstance(valor, (int, float)):
        return None
    return int(valor) if float(valor).is_integer() else float(valor)


def hoja_del_cliente(cliente_id):
    """La hoja de captura como se ve para este cliente, lista para mostrar.

    Los valores salen del archivo del cliente si ya lo generó (ahí los
    totales están calculados) y, si todavía no, de la plantilla. Encima de
    eso se ponen los valores que el contador anotó y que aún no han pasado
    por el recálculo: esos se marcan como pendientes, para que se vea que
    los totales de al lado todavía no los incluyen.
    """
    armado = mapa()
    anotados = db.listar_valores_210(cliente_id)

    archivo = archivo_cliente(cliente_id)
    plantilla = ruta_plantilla()
    if plantilla is None:
        raise SinPlantilla("No hay ninguna plantilla en plantillas/.")

    desde_archivo = archivo.exists()
    calculados = valores_calculados(archivo if desde_archivo else plantilla)

    filas = {}
    pendientes = 0

    for celda in armado["celdas"]:
        numero_fila = celda["fila"]
        if numero_fila not in filas:
            filas[numero_fila] = {
                "fila": numero_fila,
                "seccion": celda["seccion"],
                "renglon": celda["renglon"],
                "descripcion": celda["descripcion"],
                "sangria": celda["sangria"],
                "es_nota": celda["es_nota"],
                "celdas": {},
            }

        anotado = anotados.get(celda["celda"])
        calculado = _numero_o_nada(calculados.get(celda["celda"]))

        # El valor que se muestra: el anotado manda sobre lo que haya en el
        # archivo, porque es lo último que dijo el contador.
        if anotado is not None:
            mostrado = anotado["valor"]
            pendiente = (calculado != anotado["valor"])
        else:
            mostrado = calculado
            pendiente = False

        # Las casillas donde no se puede escribir y que además están
        # vacías no se mandan a la pantalla: son la mayoría y no se ven.
        if celda["tipo"] != TIPO_CAPTURA and mostrado is None:
            continue

        if pendiente:
            pendientes += 1

        filas[numero_fila]["celdas"][celda["columna"]] = {
            "celda": celda["celda"],
            "tipo": celda["tipo"],
            "editable": celda["tipo"] == TIPO_CAPTURA,
            "esperado": celda["cero_precargado"],
            "valor": mostrado,
            "anotado": anotado is not None,
            "pendiente": pendiente,
            "documento": anotado["documento"] if anotado else "",
        }

    # Se dejan afuera las filas que no dicen nada: sin descripción y sin
    # ningún valor. Son separadores en blanco de la plantilla y en pantalla
    # solo alargan la lista.
    utiles = []
    for numero_fila in sorted(filas):
        fila = filas[numero_fila]
        tiene_valor = any(
            c["valor"] is not None for c in fila["celdas"].values()
        )
        if fila["descripcion"] or tiene_valor:
            utiles.append(fila)

    return {
        "hoja": armado["hoja"],
        "plantilla": plantilla.name,
        "origen": "archivo" if desde_archivo else "plantilla",
        "generado": (
            datetime.fromtimestamp(archivo.stat().st_mtime)
            .isoformat(timespec="seconds") if desde_archivo else None
        ),
        "pendientes": pendientes,
        "filas": utiles,
    }


def guardar_valor(cliente_id, celda, valor, documento=""):
    """Guarda un valor para un cliente, después de revisar que se pueda.

    Esta es la puerta por la que entran los datos, vengan de donde vengan:
    del contador escribiéndolos a mano o de la lectura de un documento. La
    revisión es la misma para todos.
    """
    celda = str(celda).upper().strip()
    informacion = indice().get(celda)

    if informacion is None:
        raise EscrituraBloqueada(
            f"La celda {celda} no existe en la hoja de captura de la"
            f" plantilla."
        )
    if informacion["tipo"] != TIPO_CAPTURA:
        motivo = informacion["motivo"] or "no es una casilla de captura"
        raise EscrituraBloqueada(
            f"En {celda} no se puede escribir: {motivo}."
        )
    if isinstance(valor, bool) or not isinstance(valor, (int, float)):
        raise EscrituraBloqueada(
            "El valor tiene que ser un número. Para dejar una casilla en"
            " cero, escriba 0."
        )
    if valor != valor or valor in (float("inf"), float("-inf")):
        raise EscrituraBloqueada("Ese número no se puede guardar.")

    plantilla = ruta_plantilla()
    guardado = db.guardar_valor_210(
        cliente_id, celda, valor, documento,
        plantilla=plantilla.name if plantilla else "",
    )
    guardado.update(_para_pantalla(informacion))
    guardado["valor"] = db._numero(valor)
    return guardado


def listar_valores(cliente_id):
    """Los valores capturados de un cliente, con la descripción de cada uno."""
    capturados = db.listar_valores_210(cliente_id)
    catalogo = indice()

    salida = []
    for celda, guardado in capturados.items():
        informacion = catalogo.get(celda)
        fila = dict(guardado)
        # Si se anotó con otra plantilla, se avisa: la misma casilla no
        # tiene por qué significar lo mismo en dos plantillas distintas.
        actual = ruta_plantilla()
        fila["otra_plantilla"] = bool(
            guardado.get("plantilla") and actual
            and guardado["plantilla"] != actual.name
        )
        if informacion is not None:
            fila.update(_para_pantalla(informacion))
        else:
            # La plantilla cambió y esa celda ya no existe. Se muestra
            # igual, para que el contador se entere en vez de que el dato
            # desaparezca en silencio.
            fila.update({
                "descripcion": "Esta celda ya no existe en la plantilla",
                "contexto": "", "seccion": "", "renglon": "", "esperado": False,
                "columna": "", "fila": 0,
            })
        salida.append(fila)

    salida.sort(key=lambda f: (f.get("fila") or 0, f.get("columna") or ""))
    return salida


# ---------------------------------------------------------------------------
# El archivo de cada cliente
# ---------------------------------------------------------------------------


def carpeta_cliente(cliente_id):
    """datos/formularios/cliente-3/ — la carpeta propia de ese cliente."""
    return CARPETA_FORMULARIOS / f"cliente-{int(cliente_id)}"


def archivo_cliente(cliente_id):
    """La ruta del archivo generado de un cliente (exista o no todavía)."""
    return carpeta_cliente(cliente_id) / NOMBRE_ARCHIVO


def nombre_para_descargar(cliente):
    """El nombre con el que se descarga: 'Formulario 210 - Juan Pérez.xlsx'."""
    nombre = (cliente["nombre"] or "cliente").strip()
    return sanitizar_nombre(f"Formulario 210 - {nombre}.xlsx")


def eliminar_carpeta_cliente(cliente_id):
    """Borra la carpeta del cliente. Se llama al eliminar el cliente."""
    carpeta = carpeta_cliente(cliente_id)
    if carpeta.exists():
        shutil.rmtree(carpeta, ignore_errors=True)


def estado(cliente_id):
    """En qué va el formulario de un cliente: cuántos valores, qué archivo hay."""
    archivo = archivo_cliente(cliente_id)
    cuantos = len(db.listar_valores_210(cliente_id))

    if not archivo.exists():
        return {"valores": cuantos, "hay_archivo": False}

    from datetime import datetime
    generado = datetime.fromtimestamp(archivo.stat().st_mtime)
    return {
        "valores": cuantos,
        "hay_archivo": True,
        "generado": generado.isoformat(timespec="seconds"),
        "tamano": archivo.stat().st_size,
    }


def leer_totales(ruta_xlsx):
    """Lee los totales que quedaron calculados en el archivo.

    Devuelve un renglón por cada total del formulario, con su número, su
    descripción y su valor. Solo sirve si el archivo pasó por el recálculo:
    si no, los totales están sin calcular y no se devuelve ninguno.

    Los renglones de la liquidación privada (impuesto y saldos) no entran:
    ver SECCIONES_QUE_NO_SE_MUESTRAN.
    """
    # La misma lectura que ya usaron el aviso de errores y la hoja del
    # cliente: el archivo se abrió una vez y de ahí salen las tres cosas.
    hoja = valores_calculados_del_libro(ruta_xlsx).get(mapa()["hoja"], {})

    totales = []
    vistos = set()
    for celda in mapa()["celdas"]:
        if celda["columna"] != "I" or not celda["renglon"]:
            continue
        if celda["seccion"] in SECCIONES_QUE_NO_SE_MUESTRAN:
            continue
        if celda["renglon"] in vistos:
            continue
        valor = hoja.get(celda["celda"])
        if not isinstance(valor, (int, float)) or isinstance(valor, bool):
            continue
        vistos.add(celda["renglon"])
        totales.append({
            "renglon": celda["renglon"],
            "descripcion": celda["descripcion"],
            "seccion": celda["seccion"],
            "valor": int(valor) if float(valor).is_integer() else float(valor),
        })

    totales.sort(key=lambda t: int(t["renglon"]))
    return totales


def generar(cliente, con_totales=False):
    """Arma el archivo de Excel de un cliente y devuelve cómo salió.

    Siempre parte de la plantilla limpia y le escribe todos los valores
    capturados de ese cliente. El archivo anterior de ese cliente se
    reemplaza; los de los demás clientes ni se tocan.

    Por qué `con_totales` viene apagado de fábrica
    ----------------------------------------------
    Calcular los totales es pedirle a LibreOffice que abra el libro,
    recalcule las 902 fórmulas y lo vuelva a guardar. Eso cuesta cerca de
    dos segundos, y hay que sumarle revisar otra vez el archivo que
    LibreOffice reescribió entero. Es la mayor parte de lo que tarda
    generar.

    Y en la mayoría de los casos no hace falta: el archivo se entrega
    marcado para que **Excel calcule los totales solo al abrirlo** (lo hace
    `_marcar_para_recalcular` en escribir_210.py). El contador descarga su
    archivo y ve sus totales, igual que siempre.

    Lo único que se gana calculando aquí es poder mostrar los totales
    DENTRO de Tax-i, en su propia tabla. Eso es útil, pero es una consulta,
    no parte de guardar: por eso ahora se pide aparte, con el botón
    «Actualizar totales».

    Lo que NO cambia: la verificación de las 902 fórmulas contra la
    plantilla original se hace siempre, en las dos formas. Es la que
    protege el archivo del contador y no tiene interruptor.
    """
    plantilla = ruta_plantilla()
    if plantilla is None:
        raise SinPlantilla(
            "No hay ninguna plantilla en la carpeta plantillas/."
        )

    cliente_id = cliente["id"]
    valores = db.listar_valores_210(cliente_id)

    escritor = EscritorPlantilla(
        plantilla,
        nombre_salida=NOMBRE_ARCHIVO,
        carpeta_trabajo=carpeta_cliente(cliente_id),
    )
    for celda, guardado in valores.items():
        escritor.escribir(
            celda,
            guardado["valor"],
            documento=guardado["documento"] or "digitado por el contador",
        )

    ruta, ruta_bitacora = escritor.guardar(recalcular_totales=con_totales)

    informe = {
        "archivo": str(ruta.relative_to(RAIZ)),
        "nombre_descarga": nombre_para_descargar(cliente),
        "valores_escritos": len(valores),
        "verificacion": escritor.informe_verificacion,
        "recalculo": escritor.informe_recalculo,
        "bitacora": str(ruta_bitacora.relative_to(RAIZ)),
        # Si no se calcularon los totales, la pantalla lo dice y ofrece el
        # botón para hacerlo. No es un error: el archivo está completo.
        "con_totales": bool(con_totales),
        "totales": [],
    }
    if escritor.informe_recalculo and escritor.informe_recalculo["recalculado"]:
        informe["totales"] = leer_totales(ruta)

    return informe
