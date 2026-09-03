"""
Arma un reporte de exógena inventado, con la estructura real de la DIAN.

Sirve para dos cosas: para probar el programa de punta a punta sin usar
el archivo de nadie, y para que el contador vea cómo se comporta Tax-i
antes de cargarle el de un cliente de verdad.

TODO lo de aquí es inventado: el contribuyente, las cifras, los números
de cuenta y las empresas pequeñas. Los NIT de los bancos y de la DIAN sí
son los públicos de esas instituciones, porque si no el cruce no se
parecería en nada al de verdad.

Los textos de la columna «Uso declaración Sugerida» están copiados de un
reporte real, palabra por palabra. Eso NO es un adorno: son de la DIAN,
y las decisiones que el contador toma mirándolos dependen de que digan
exactamente lo que ella escribió. Aquí no se reescribe ninguno.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

# Los tres avisos legales, textuales.
AVISOS = (
    "ADVERTENCIA: Esta información corresponde a la fecha de corte del"
    " proceso y puede estar sujeta a cambios de los terceros que la reportan"
    " conforme a las modificaciones o adiciones del informante",

    "Si esta información presenta inconsistencias, debe comunicarse o acudir"
    " a la persona natural o jurídica que entrega la información. Sus datos"
    " de identificación están en la columna 'persona que reporta'. ",

    "IMPORTANTE: Para cumplir con su obligación de declarar, la Información"
    " Exógena Tributaria NO ES INDISPENSABLE y NO REEMPLAZA la información"
    " de su realidad económica, ni lo exonera de declarar los valores totales"
    " que correspondan y que son de su conocimiento exclusivo.",
)

ENCABEZADOS = (
    "NIT", "Nombre / Razón Social", "NIT",
    "Nombre/Razón Social reportada por el tercero",
    "Detalle", "Valor", "Uso declaración Sugerida", "Información  Adicional ",
)

TOPES = (
    ("Tope 1 - Ingresos", 93520000),
    ("Tope 2 - Patrimonio", 210400000),
    ("Tope 3 - Consumo TC", 22400000),
    ("Tope 4 - Movimiento", 96300000),
    ("Tope 5 - Compras", 14200000),
)

# Los renglones y los topes se citan con el texto exacto de la DIAN.
R29_R30 = ("Tope 2: Patrimonio | R29 Patrimonio Bruto (si el saldo es"
           " positivo)| R30 Deudas  (si el saldo es negativo)")
R32 = ("Tope 1: Ingresos brutos | R32 Ingresos brutos por rentas de trabajo"
       " (art. 103 E.T.)")
R58_R59 = ("Tope 1: Ingresos brutos | R58 Ingresos brutos por rentas de"
           " capital | R59 Ingresos no constitutivos por rentas de capital")
R132 = "R132 Retenciones año gravable a declarar"
R30 = "R30 Deudas"
R74 = "Tope 1: Ingresos brutos | R74 Ingresos brutos rentas no laborales "
NO_CONSTITUTIVOS_TRABAJO = ("Ingresos no constitutivos de renta | Asignación"
                            " según el tipo de renta: R33 (Trabajo) / R59"
                            " (Capital) / R100 (Pensiones)")
NO_CONSTITUTIVOS_PENSION = ("Ingresos no constitutivos de renta | Asignación"
                            " según el tipo de renta: R33 (Trabajo) / R59"
                            " (Capital) / R76 (No laborales)")
CESANTIAS_EMPLEADOR = (
    "Tope 1: Ingresos brutos | Tope 2: Patrimonio | R29 Patrimonio Bruto |"
    " R36 Otras rentas exentas\nNota: este valor puede ser reportado por el"
    " empleador y el fondo de cesantías, el sistema toma un solo registro"
    " para evitar duplicidad.")
CESANTIAS_FONDO = (
    "Tope 1: Ingresos brutos | R29 Patrimonio Bruto |  R32 Ingresos brutos"
    " por rentas de trabajo (art. 103 E.T.) | Según el tipo de renta se"
    " incluye en los cálculos de R36 Otras rentas exentas (laborales) o"
    " Otras deducciones imputables para R51 (Honorarios), R67 (Capital) o"
    " R84 (No laborales).\nNota: este valor puede ser reportado por el"
    " empleador y el fondo de cesantías, el sistema toma un solo registro"
    " para evitar duplicidad.")
PREDIAL = ("Tope 2: Patrimonio | R29 Patrimonio Bruto | Se toma el mayor"
           " valor informado entre el avalúo y la base del impuesto predial")
DOCUMENTO_SOPORTE = (
    'Son los ingresos de sus ventas/servicios que quedaron registrados a'
    ' través de "documentos soporte" emitidos por sus compradores. El sistema'
    ' compara el total de estos documentos con el de su Información Exógena y'
    ' selecciona el mayor para el Tope 1: Ingresos brutos.')

# Quién le reportó qué. (NIT, nombre, detalle, valor, uso, información
# adicional). Los tres primeros terceros son los mismos de los documentos
# que ya tiene el cliente de prueba, para que el cruce se vea funcionando.
FILAS = (
    (900123456, "COMERCIALIZADORA EL ROBLE S.A.S.",
     "Pagos por salarios (Concepto: 2276)", 84600000, R32, ""),
    (900123456, "COMERCIALIZADORA EL ROBLE S.A.S.",
     "Cesantías consignadas al fondo de cesantías (Concepto: 2276)",
     7050000, CESANTIAS_EMPLEADOR, ""),
    (900123456, "COMERCIALIZADORA EL ROBLE S.A.S.",
     "Aportes obligatorios a salud a cargo Trabajador (Concepto: 2276)",
     3384000, NO_CONSTITUTIVOS_TRABAJO, ""),
    (900123456, "COMERCIALIZADORA EL ROBLE S.A.S.",
     "Aporte obligatorio fondos pensiones y solidaridad a cargo del"
     " trabajador (Concepto: 2276)", 3384000, NO_CONSTITUTIVOS_PENSION, ""),
    (900123456, "COMERCIALIZADORA EL ROBLE S.A.S.",
     "Retención en la fuente practicada (Concepto: 2276)", 4120000, R132, ""),

    (860034313, "BANCO DAVIVIENDA S.A.",
     "Saldo cuentas bancarias (Titular Principal)", 18750000, R29_R30,
     "Número de Cuenta / Documento: 008744552211 | Concepto Códigos"
     " Tributaria: *1* Cuenta de ahorro"),
    (860034313, "BANCO DAVIVIENDA S.A.",
     "CDT Rendimientos Pagados (Informado principal) (Concepto: 1020)",
     620000, R58_R59, "Número de Cuenta / Documento: 008744552211"),
    (860034313, "BANCO DAVIVIENDA S.A.",
     "CDT Retención prácticada (Concepto: 1020)", 62000, R132,
     "Número de Cuenta / Documento: 008744552211"),
    (860034313, "BANCO DAVIVIENDA S.A.",
     "Total consumos o gastos con tarjeta Crédito o Débito (Concepto: 1023)",
     22400000, "Tope 3: Consumos TC",
     "Número de Cuenta / Documento: 4417 | Clase de Tarjeta: *1* Tarjeta de"
     " Credito Principal"),
    (860034313, "BANCO DAVIVIENDA S.A.",
     "Valor total de los movimientos en cuentas corrientes y de ahorros "
     " (Titular Principal)", 96300000, "Tope 4: Consignaciones e inversiones",
     "Número de Cuenta / Documento: 008744552211"),

    (890903938, "BANCOLOMBIA S.A.",
     "Cuentas por pagar de clientes (Concepto: 1315)", 96400000, R30, ""),

    (800229739, "FONDO DE CESANTIAS PROTECCION",
     "Valor total de las cesantías abonadas en el periodo.  (Formato del"
     " fondo de cesantías).  Empleado", 7049100, CESANTIAS_FONDO,
     "Tipo de Afiliado: *1* Trabajador | Tipo de Aportante: ** "),

    (890900111, "MUNICIPIO DE VILLA DEL ROBLE",
     "Valor avalúo catastral (Concepto: 1476)", 185000000, PREDIAL,
     "Porcentaje de Participación: 100.00 | Matricula: 0 | Número"
     " Propietarios: 1"),

    (901555222, "TALLERES Y SERVICIOS DEL SUR S.A.S.",
     "Ingresos Documentos soporte de adquisiciones no obligados a expedir"
     " factura de venta o documento equivalente", 1250000, DOCUMENTO_SOPORTE,
     "Documento: DSE0117 | CUDS:"
     " c4a1b8d6f30295e7148b0c6d3f9a527e04b1c8d6a3f70925e18b4c0d6f3a92750"),
    (901555222, "TALLERES Y SERVICIOS DEL SUR S.A.S.",
     "Otros ingresos (Concepto: 5016)", 1250000, R74, ""),

    (800123987, "COLSANITAS MEDICINA PREPAGADA",
     "Pagos por medicina prepagada", 5400000, "", ""),

    (800197268, "U.A.E. DIRECCION DE IMPUESTOS Y ADUANAS NACIONALES",
     "Suma valor total facturas tras ajustes por notas", 14200000,
     "Tope 5: Compras registradas en Factura Electrónica", ""),
)


def escribir(ruta, nombre="PEDRO RUIZ MARTINEZ", cedula="79845641",
             anio="2025", corte=None):
    """Escribe el reporte inventado y devuelve la ruta.

    La estructura es la del archivo que baja del portal de la DIAN: la
    cabecera con los avisos, los encabezados en la fila 14, los cinco
    topes y después los registros. El lector no depende de esos números
    de fila —busca «Uso declaración Sugerida»— pero el archivo sí tiene
    que parecerse al de verdad para que la prueba sirva de algo.
    """
    corte = corte or datetime(2026, 8, 26)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Reporte"

    hoja["C1"] = "Consulta de Información reportada por terceros"
    hoja["A2"] = AVISOS[0]
    hoja["G2"] = "Fecha   Reporte:"
    hoja["H2"] = corte.replace(hour=21, minute=57, second=56)
    hoja["A3"] = "Fecha corte del proceso: "
    hoja["C3"] = corte
    hoja["A4"] = "Año al que se refiere la consulta:"
    hoja["C4"] = anio
    hoja["A5"] = "Identificación del consultante"
    hoja["A6"] = "Tipo de documento:"
    hoja["C6"] = "C. C."
    hoja["A7"] = "Identificación:"
    # Como texto, que es como lo escribe la DIAN para el contribuyente.
    hoja["C7"] = str(cedula)
    hoja["A8"] = "Nombres / Razón social:"
    hoja["C8"] = nombre
    hoja["A10"] = AVISOS[1]
    hoja["A11"] = AVISOS[2]
    hoja["A13"] = "Persona que reporta"
    hoja["C13"] = "Información reportada"

    for columna, titulo in enumerate(ENCABEZADOS, start=1):
        hoja.cell(row=14, column=columna, value=titulo)

    for posicion, (etiqueta, valor) in enumerate(TOPES):
        hoja.cell(row=15 + posicion, column=5, value=etiqueta)
        hoja.cell(row=15 + posicion, column=6, value=valor)

    for posicion, fila in enumerate(FILAS):
        nit, quien, detalle, valor, uso, adicional = fila
        numero = 20 + posicion
        # El NIT de quien reporta va como NÚMERO y el del contribuyente
        # como texto: así vienen mezclados en el archivo de verdad.
        hoja.cell(row=numero, column=1, value=nit)
        hoja.cell(row=numero, column=2, value=quien)
        hoja.cell(row=numero, column=3, value=str(cedula))
        hoja.cell(row=numero, column=4, value=nombre)
        hoja.cell(row=numero, column=5, value=detalle)
        hoja.cell(row=numero, column=6, value=valor)
        if uso:
            hoja.cell(row=numero, column=7, value=uso)
        if adicional:
            hoja.cell(row=numero, column=8, value=adicional)

    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    libro.save(ruta)
    return ruta
