"""
Prueba del Paso 2: escribir en la plantilla sin dañarla.

Se corre así, desde la carpeta del proyecto:

    .venv/bin/python pruebas/probar_escritura_210.py

No usa ninguna librería de pruebas: es un programa normal que hace cosas y
dice si salieron bien. Imprime una línea por comprobación. Al final dice
cuántas pasaron.

Lo que comprueba, en dos partes:

  A. Que se NIEGUE a escribir donde no debe (fórmula, otra hoja, otra
     columna, texto, celda combinada, fila fuera de la tabla).
  B. Que cuando sí escribe, el archivo quede sano: el original intacto, las
     902 fórmulas iguales, las 22 imágenes en su lugar y los valores
     escritos donde se pidió.
  C. Que la verificación posterior (regla 6) se corra sola al guardar, que
     detecte un archivo dañado, y que cuando lo detecta borre el archivo en
     vez de entregarlo.
  D. Que el recálculo con LibreOffice deje los totales listos para leer, y
     que si LibreOffice no está o se demora, el archivo se entregue igual
     con un aviso en vez de romperse.
"""

import hashlib
import json
import re
import sys
import zipfile
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from openpyxl import Workbook, load_workbook  # noqa: E402

import app.escribir_210 as escribir_210  # noqa: E402
import app.recalcular as recalcular_lo  # noqa: E402
from app.escribir_210 import (  # noqa: E402
    EscrituraBloqueada,
    EscritorPlantilla,
    VerificacionFallida,
    verificar_contra_original,
)
from app.plantilla_210 import contar_formulas_del_libro, es_formula  # noqa: E402

PLANTILLAS = RAIZ / "plantillas"

resultados = []


def comprobar(descripcion, condicion, detalle=""):
    """Anota una comprobación y la imprime."""
    resultados.append(bool(condicion))
    marca = "OK  " if condicion else "FALLA"
    linea = f"  {marca}  {descripcion}"
    if detalle:
        linea += f"  [{detalle}]"
    print(linea)


def se_niega(descripcion, funcion):
    """Comprueba que una escritura sea rechazada, y muestra el motivo."""
    try:
        funcion()
    except EscrituraBloqueada as error:
        comprobar(descripcion, True, str(error)[:70])
        return
    comprobar(descripcion, False, "NO se negó: la escribió")


def huella(ruta):
    return hashlib.md5(Path(ruta).read_bytes()).hexdigest()


def buscar_plantilla():
    encontradas = sorted(
        p for p in PLANTILLAS.glob("*.xlsx") if not p.name.startswith("~$")
    )
    if not encontradas:
        print("No hay ninguna plantilla .xlsx en plantillas/. No se puede probar.")
        sys.exit(1)
    return encontradas[0]


def main():
    plantilla = buscar_plantilla()
    print(f"Plantilla: {plantilla.name}")

    huella_antes = huella(plantilla)
    formulas_antes = contar_formulas_del_libro(plantilla)
    partes_antes = set(zipfile.ZipFile(plantilla).namelist())

    # ------------------------------------------------------------------
    print("\nA. Escrituras que se deben rechazar")
    # ------------------------------------------------------------------
    escritor = EscritorPlantilla(plantilla, nombre_salida="prueba_paso2.xlsx")

    se_niega(
        "celda con fórmula (I28 = IF(ROUND(H30*1%...)))",
        lambda: escritor.escribir("I28", 123, documento="prueba"),
    )
    se_niega(
        "celda con fórmula escrita a mano (G58)",
        lambda: escritor.escribir("G58", 123, documento="prueba"),
    )
    se_niega(
        "otra hoja (Formulario 210)",
        lambda: escritor.escribir(
            "G32", 123, documento="prueba", hoja="Formulario 210"
        ),
    )
    se_niega(
        "hoja de tablas estructurales (Tablas de impuestos)",
        lambda: escritor.escribir(
            "G32", 123, documento="prueba", hoja="Tablas de impuestos"
        ),
    )
    se_niega(
        "columna que no es de valores (F32, la descripción)",
        lambda: escritor.escribir("F32", 123, documento="prueba"),
    )
    se_niega(
        "texto en vez de número",
        lambda: escritor.escribir("G32", "1.500.000", documento="prueba"),
    )
    se_niega(
        "vacío en vez de 0",
        lambda: escritor.escribir("G32", None, documento="prueba"),
    )
    se_niega(
        "celda combinada que no es la esquina (H45)",
        lambda: escritor.escribir("H45", 123, documento="prueba"),
    )
    se_niega(
        "fila fuera de la tabla (G5000)",
        lambda: escritor.escribir("G5000", 123, documento="prueba"),
    )
    se_niega(
        "coordenada inventada (ZZ)",
        lambda: escritor.escribir("ZZ", 123, documento="prueba"),
    )

    comprobar(
        "después de 10 rechazos no quedó ningún cambio anotado",
        escritor.cambios == {} and escritor.bitacora == [],
        f"cambios={len(escritor.cambios)}",
    )

    # ------------------------------------------------------------------
    print("\nB. Escrituras válidas")
    # ------------------------------------------------------------------
    escritor.escribir("G32", 1500000, documento="factura_electronica.pdf")
    escritor.escribir("H104", 2300000, documento="certificado_banco.pdf")
    escritor.escribir("I484", 0, documento="digitado por el contador")
    # La misma celda dos veces: debe quedar el último valor, y la bitácora
    # debe registrar los dos movimientos.
    escritor.escribir("G32", 1750000, documento="factura_electronica.pdf (v2)")

    # Sin recálculo a propósito: así se comprueba que la escritura
    # quirúrgica no tocó nada. El recálculo se prueba aparte, en la
    # sección K, porque LibreOffice reescribe el archivo entero.
    ruta_salida, ruta_bitacora = escritor.guardar(recalcular_totales=False)
    print(f"  archivo generado: {ruta_salida.relative_to(RAIZ)}")

    # ------------------------------------------------------------------
    print("\nC. El original quedó intacto")
    # ------------------------------------------------------------------
    comprobar(
        "la plantilla original no cambió (misma huella MD5)",
        huella(plantilla) == huella_antes,
    )
    comprobar(
        "el archivo generado es otro archivo",
        ruta_salida.resolve() != plantilla.resolve(),
    )

    # ------------------------------------------------------------------
    print("\nD. El archivo generado quedó sano")
    # ------------------------------------------------------------------
    partes_despues = set(zipfile.ZipFile(ruta_salida).namelist())
    perdidas = partes_antes - partes_despues
    comprobar(
        f"no se perdió ninguna de las {len(partes_antes)} partes internas",
        not perdidas,
        f"perdidas: {sorted(perdidas)[:3]}" if perdidas else "",
    )

    imagenes_antes = [p for p in partes_antes if p.startswith("xl/media/")]
    imagenes_despues = [p for p in partes_despues if p.startswith("xl/media/")]
    comprobar(
        f"las {len(imagenes_antes)} imágenes siguen ahí (incluidas las de Copyright)",
        len(imagenes_antes) == len(imagenes_despues),
        f"{len(imagenes_despues)} encontradas",
    )

    formulas_despues = contar_formulas_del_libro(ruta_salida)
    comprobar(
        f"las {sum(formulas_antes.values())} fórmulas del libro siguen ahí",
        formulas_antes == formulas_despues,
        f"antes {sum(formulas_antes.values())} / después {sum(formulas_despues.values())}",
    )

    libro_antes = load_workbook(plantilla, data_only=False)
    libro_despues = load_workbook(ruta_salida, data_only=False)
    comprobar(
        "las 15 hojas siguen con el mismo nombre y en el mismo orden",
        libro_antes.sheetnames == libro_despues.sheetnames,
    )

    distintas = []
    for nombre in libro_antes.sheetnames:
        hoja_a, hoja_d = libro_antes[nombre], libro_despues[nombre]
        for fila in hoja_a.iter_rows():
            for celda in fila:
                if es_formula(celda.value):
                    if hoja_d[celda.coordinate].value != celda.value:
                        distintas.append(f"{nombre}!{celda.coordinate}")
    comprobar(
        "ni una sola fórmula cambió de texto",
        not distintas,
        f"cambiadas: {distintas[:3]}" if distintas else "",
    )

    # ------------------------------------------------------------------
    print("\nE. Los valores quedaron donde se pidió")
    # ------------------------------------------------------------------
    hoja = libro_despues["Detalle renglón 210"]
    comprobar("G32 quedó en 1750000 (el último valor escrito)",
              hoja["G32"].value == 1750000, repr(hoja["G32"].value))
    comprobar("H104 quedó en 2300000",
              hoja["H104"].value == 2300000, repr(hoja["H104"].value))
    comprobar("I484 quedó en 0 (limpiar es escribir 0, no vaciar)",
              hoja["I484"].value == 0, repr(hoja["I484"].value))

    hoja_original = libro_antes["Detalle renglón 210"]
    intactas = ["G44", "H105", "I665", "G115", "H204"]
    iguales = all(
        hoja[c].value == hoja_original[c].value for c in intactas
    )
    comprobar("las celdas vecinas que no se tocaron siguen igual", iguales,
              ", ".join(intactas))

    formato_igual = all(
        hoja[c].number_format == hoja_original[c].number_format
        for c in ("G32", "H104", "I484")
    )
    comprobar("el formato de número de las celdas escritas no cambió",
              formato_igual)

    xml = zipfile.ZipFile(ruta_salida).read("xl/workbook.xml").decode("utf-8")
    comprobar("el libro quedó marcado para recalcular al abrirlo en Excel",
              "fullCalcOnLoad" in xml)

    # ------------------------------------------------------------------
    print("\nF. La bitácora")
    # ------------------------------------------------------------------
    datos = json.loads(ruta_bitacora.read_text(encoding="utf-8"))
    cambios = datos["cambios"]
    comprobar("quedaron los 4 movimientos, no 3", len(cambios) == 4,
              f"{len(cambios)} movimientos")
    primero = cambios[0]
    comprobar(
        "cada movimiento trae hoja, celda, antes, después, documento y hora",
        all(k in primero for k in
            ("hoja", "celda", "valor_anterior", "valor_nuevo", "documento",
             "fecha_hora")),
    )
    comprobar("el primer cambio de G32 registra el valor anterior (0)",
              primero["celda"] == "G32" and primero["valor_anterior"] == 0,
              f"anterior={primero['valor_anterior']!r}")
    comprobar("el segundo cambio de G32 registra como anterior el 1500000",
              cambios[3]["valor_anterior"] == 1500000,
              f"anterior={cambios[3]['valor_anterior']!r}")
    comprobar("la hora quedó registrada",
              bool(re.match(r"\d{4}-\d{2}-\d{2}T", primero["fecha_hora"])),
              primero["fecha_hora"])

    # ------------------------------------------------------------------
    print("\nG. No se puede seguir escribiendo después de guardar")
    # ------------------------------------------------------------------
    se_niega(
        "escribir después de guardar",
        lambda: escritor.escribir("G44", 1, documento="prueba"),
    )

    # ------------------------------------------------------------------
    print("\nH. La verificación posterior (regla 6) se corrió sola al guardar")
    # ------------------------------------------------------------------
    informe = escritor.informe_verificacion
    comprobar("guardar() dejó un informe de verificación", informe is not None)
    comprobar(f"comparó las {sum(formulas_antes.values())} fórmulas",
              informe["formulas_comparadas"] == sum(formulas_antes.values()),
              f"{informe['formulas_comparadas']} comparadas")
    comprobar("no encontró ninguna fórmula distinta",
              informe["formulas_distintas"] == 0)
    comprobar("solo cambiaron la hoja de captura y workbook.xml",
              sorted(informe["partes_modificadas"]) ==
              ["xl/workbook.xml", "xl/worksheets/sheet5.xml"],
              ", ".join(informe["partes_modificadas"]))
    comprobar("el informe quedó guardado dentro de la bitácora",
              "verificacion" in json.loads(
                  ruta_bitacora.read_text(encoding="utf-8")))

    # ------------------------------------------------------------------
    print("\nI. La verificación detecta un archivo dañado")
    # ------------------------------------------------------------------
    # Se fabrican a propósito los dos daños que más miedo dan, y se
    # comprueba que la verificación los cache. Estos archivos se hacen en
    # una carpeta aparte y no se entregan a nadie.
    carpeta_daños = RAIZ / "datos" / "trabajo" / "dañados_de_prueba"
    carpeta_daños.mkdir(parents=True, exist_ok=True)

    # Daño 1: guardar con data_only=True. Es la catástrofe que la regla 2
    # prohíbe: reemplaza todas las fórmulas por su último resultado.
    sin_formulas = carpeta_daños / "sin_formulas.xlsx"
    libro_roto = load_workbook(plantilla, data_only=True)
    libro_roto.save(sin_formulas)
    try:
        verificar_contra_original(plantilla, sin_formulas)
        comprobar("detecta un archivo con las fórmulas destruidas", False,
                  "no la detectó")
    except VerificacionFallida as error:
        comprobar("detecta un archivo con las fórmulas destruidas", True,
                  str(error).splitlines()[1].strip()[:60])

    # Daño 2: abrir y guardar con openpyxl sin tocar nada. Las fórmulas
    # sobreviven, pero se pierden las imágenes. Es el camino que se
    # descartó, y la verificación tiene que verlo.
    sin_imagenes = carpeta_daños / "sin_imagenes.xlsx"
    load_workbook(plantilla).save(sin_imagenes)
    try:
        verificar_contra_original(plantilla, sin_imagenes)
        comprobar("detecta un archivo que perdió las imágenes", False,
                  "no la detectó")
    except VerificacionFallida as error:
        comprobar("detecta un archivo que perdió las imágenes", True,
                  str(error).splitlines()[1].strip()[:60])

    # ------------------------------------------------------------------
    print("\nJ. Si la verificación falla, el archivo se descarta")
    # ------------------------------------------------------------------
    # Se reemplaza a propósito la parte que escribe por una que daña el
    # archivo, para ver qué hace guardar() cuando la verificación falla.
    escritor_saboteado = EscritorPlantilla(
        plantilla, nombre_salida="prueba_paso3_saboteada.xlsx"
    )
    escritor_saboteado.escribir("G32", 999, documento="prueba de sabotaje")
    ruta_saboteada = escritor_saboteado.ruta_copia

    original_de_verdad = escribir_210._aplicar_cambios_al_zip

    def sabotear(ruta_xlsx, cambios):
        """Hace justo lo que nunca se debe hacer: destruir las fórmulas."""
        load_workbook(ruta_xlsx, data_only=True).save(ruta_xlsx)

    escribir_210._aplicar_cambios_al_zip = sabotear
    try:
        escritor_saboteado.guardar()
        comprobar("guardar() se niega a entregar un archivo dañado", False,
                  "lo entregó")
    except VerificacionFallida as error:
        comprobar("guardar() se niega a entregar un archivo dañado", True,
                  str(error).splitlines()[1].strip()[:55])
    finally:
        escribir_210._aplicar_cambios_al_zip = original_de_verdad

    comprobar("el archivo dañado se borró, no quedó dando vueltas",
              not ruta_saboteada.exists(), str(ruta_saboteada.name))
    comprobar("no se escribió bitácora para un archivo que no se entregó",
              not Path(str(ruta_saboteada) + ".bitacora.json").exists())

    # Los archivos dañados a propósito no se dejan por ahí: alguien los
    # podría confundir con un resultado bueno.
    for basura in carpeta_daños.glob("*.xlsx"):
        basura.unlink()
    carpeta_daños.rmdir()

    # ------------------------------------------------------------------
    print("\nK. El recálculo con LibreOffice")
    # ------------------------------------------------------------------
    hay_libreoffice = recalcular_lo.buscar_libreoffice()
    print(f"  LibreOffice: {hay_libreoffice or 'no está instalado'}")

    escritor_lo = EscritorPlantilla(plantilla, nombre_salida="prueba_paso4.xlsx")
    escritor_lo.escribir("G32", 1000000, documento="factura_1.pdf")
    escritor_lo.escribir("G34", 500000, documento="factura_2.pdf")
    ruta_lo, bitacora_lo = escritor_lo.guardar()
    informe_lo = escritor_lo.informe_recalculo

    comprobar("guardar() dejó un informe del recálculo", informe_lo is not None)

    if hay_libreoffice:
        comprobar("LibreOffice recalculó el libro",
                  informe_lo["recalculado"], informe_lo["motivo"][:55])
        comprobar("tardó menos que el tiempo límite de 120 segundos",
                  informe_lo["segundos"] < 120,
                  f"{informe_lo['segundos']} segundos")

        calculado = load_workbook(ruta_lo, data_only=True)["Detalle renglón 210"]
        comprobar("H30 (=SUM(G32:G40)) ya trae el total: 1.500.000",
                  calculado["H30"].value == 1500000,
                  repr(calculado["H30"].value))
        comprobar("I28 calculó solo el 1% de ese total: 15.000",
                  calculado["I28"].value == 15000,
                  repr(calculado["I28"].value))
        comprobar("un total que no se tocó sigue igual (I42)",
                  calculado["I42"].value == 102003000,
                  repr(calculado["I42"].value))

        comprobar("no quedó ninguna celda con #REF!, #VALUE! y parecidos",
                  recalcular_lo.celdas_con_error(ruta_lo) == [],
                  f"{len(recalcular_lo.celdas_con_error(ruta_lo))} errores")

        formulas_lo = contar_formulas_del_libro(ruta_lo)
        comprobar("después de pasar por LibreOffice siguen las 902 fórmulas",
                  formulas_lo == formulas_antes,
                  f"{sum(formulas_lo.values())} fórmulas")

        partes_lo = set(zipfile.ZipFile(ruta_lo).namelist())
        comprobar("LibreOffice no perdió ninguna parte interna",
                  not (partes_antes - partes_lo),
                  f"perdidas: {len(partes_antes - partes_lo)}")
        comprobar("las 22 imágenes sobrevivieron al recálculo",
                  len([p for p in partes_lo if p.startswith("xl/media/")]) == 22)
        comprobar("la verificación se volvió a correr después de recalcular",
                  escritor_lo.informe_verificacion["formulas_distintas"] == 0)
        comprobar("el recálculo quedó anotado en la bitácora",
                  json.loads(bitacora_lo.read_text(encoding="utf-8"))
                  ["recalculo"]["recalculado"] is True)
    else:
        comprobar("sin LibreOffice, el archivo se entrega igual",
                  ruta_lo.exists())
        comprobar("y el aviso explica que los totales se ven en Excel",
                  "Excel" in informe_lo["motivo"], informe_lo["motivo"][:55])

    # ------------------------------------------------------------------
    print("\nL. Cuando LibreOffice no está o falla")
    # ------------------------------------------------------------------
    # Se finge que LibreOffice no está instalado, que es como va a estar el
    # computador del contador hasta que alguien se lo instale.
    buscar_de_verdad = recalcular_lo.buscar_libreoffice
    recalcular_lo.buscar_libreoffice = lambda: None
    try:
        escritor_sin_lo = EscritorPlantilla(
            plantilla, nombre_salida="prueba_paso4_sin_libreoffice.xlsx"
        )
        escritor_sin_lo.escribir("G32", 777000, documento="prueba sin LibreOffice")
        ruta_sin_lo, _ = escritor_sin_lo.guardar()
        informe_sin = escritor_sin_lo.informe_recalculo
        comprobar("sin LibreOffice el archivo se entrega igual, no se rompe",
                  ruta_sin_lo.exists())
        comprobar("el informe dice que no se recalculó",
                  informe_sin["recalculado"] is False)
        comprobar("el aviso le explica al contador qué hacer",
                  "Excel" in informe_sin["motivo"], informe_sin["motivo"][:60])
        comprobar("y el archivo entregado sigue teniendo sus 902 fórmulas",
                  contar_formulas_del_libro(ruta_sin_lo) == formulas_antes)
    finally:
        recalcular_lo.buscar_libreoffice = buscar_de_verdad

    # Tiempo límite: se le da un plazo imposible para ver qué hace.
    if hay_libreoffice:
        informe_corto = recalcular_lo.recalcular(ruta_lo, tiempo_limite=0.05)
        comprobar("si se pasa del tiempo límite, avisa en vez de colgarse",
                  informe_corto["recalculado"] is False
                  and "demor" in informe_corto["motivo"],
                  informe_corto["motivo"][:55])
        comprobar("y el archivo queda intacto tras el intento fallido",
                  contar_formulas_del_libro(ruta_lo) == formulas_antes)

    # ------------------------------------------------------------------
    print("\nM. Los errores dentro de las celdas (#REF!, #VALUE!...)")
    # ------------------------------------------------------------------
    if hay_libreoffice:
        # Un libro chiquito con dos fórmulas que fallan de verdad. Lo
        # calcula LibreOffice, no lo fingimos nosotros.
        libro_malo = RAIZ / "datos" / "trabajo" / "con_errores.xlsx"
        wb_malo = Workbook()
        hoja_mala = wb_malo.active
        hoja_mala["A1"] = 0
        hoja_mala["A2"] = "=1/A1"                    # da #DIV/0!
        hoja_mala["A3"] = "=FUNCION_QUE_NO_EXISTE(1)"  # da #NAME?
        wb_malo.save(libro_malo)

        comprobar("antes de calcular no se ve ningún error",
                  recalcular_lo.celdas_con_error(libro_malo) == [])
        recalcular_lo.recalcular(libro_malo)
        errores = recalcular_lo.celdas_con_error(libro_malo)
        comprobar("después de calcular, encuentra el #DIV/0! y el #NAME?",
                  len(errores) == 2, ", ".join(errores))
        libro_malo.unlink()

    # Y que guardar() descarte el archivo si aparecen errores nuevos.
    errores_de_verdad = escribir_210.celdas_con_error

    def fingir_errores(ruta):
        """Dice que la copia quedó con un #REF!, pero el original no."""
        if "trabajo" in str(ruta):
            return ["Detalle renglón 210!I42=#REF!"]
        return []

    escritor_con_error = EscritorPlantilla(
        plantilla, nombre_salida="prueba_paso4_con_error.xlsx"
    )
    escritor_con_error.escribir("G32", 1, documento="prueba de error")
    ruta_con_error = escritor_con_error.ruta_copia
    escribir_210.celdas_con_error = fingir_errores
    try:
        escritor_con_error.guardar()
        comprobar("si aparece un #REF! nuevo, no se entrega el archivo", False,
                  "lo entregó")
    except VerificacionFallida as error:
        comprobar("si aparece un #REF! nuevo, no se entrega el archivo", True,
                  str(error).splitlines()[0][:58])
    finally:
        escribir_210.celdas_con_error = errores_de_verdad

    comprobar("el archivo con errores se borró",
              not ruta_con_error.exists(), ruta_con_error.name)

    total = len(resultados)
    buenas = sum(resultados)
    print(f"\n{buenas} de {total} comprobaciones pasaron.")
    if buenas != total:
        print("HAY FALLAS. No seguir al Paso 3 sin arreglarlas.")
        return 1
    print("Todo bien.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
