"""
Prueba del lector de información exógena.

Se corre así, desde la carpeta del proyecto:

    .venv/bin/python pruebas/probar_exogena.py

Trabaja sobre pruebas/ejemplos/reporteExogena2025_EJEMPLO.xlsx, que
tiene la estructura real de la DIAN con datos INVENTADOS. Ningún dato
de un cliente de verdad entra aquí.

Aquí NO hay ninguna llamada a un modelo de IA, y no hace falta: el
reporte de la DIAN es una tabla bien formada y se lee con código. Esta
prueba pasa igual con IA_PROVEEDOR=ninguno, que es el modo de fábrica.

Lo que comprueba:

  A. Encuentra la fila de encabezados SIN tenerla escrita a mano, y
     falla claro cuando falta la columna «Uso declaración Sugerida».
  B. Lee la cabecera: año, documento, nombre y fecha de corte.
  C. Los tres avisos legales salen TEXTUALES, sin resumir.
  D. Lee los cinco topes, que son resumen y no renglones.
  E. Lee las 36 filas de datos con sus NIT como texto.
  F. Saca los códigos de concepto que vienen dentro del detalle.
  G. Saca los códigos R, incluidas las filas que traen varios.
  H. Marca las 8 filas que requieren decisión del contador, con las
     opciones palabra por palabra.
  I. Detecta los posibles duplicados, entre ellos las cesantías que
     reportan el empleador y el fondo.
  J. Le pone nombre a los quince renglones sin inventarse ninguno.
"""

import shutil
import sys
import tempfile
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from openpyxl import load_workbook  # noqa: E402

from app import exogena  # noqa: E402

EJEMPLO = RAIZ / "pruebas" / "ejemplos" / "reporteExogena2025_EJEMPLO.xlsx"

fallos = []
TOTAL = [0]


def revisar(que, condicion, detalle=""):
    TOTAL[0] += 1
    print(("  OK    " if condicion else "  FALLA ") + que
          + ("  [%s]" % str(detalle)[:70] if detalle else ""))
    if not condicion:
        fallos.append(que)


print("=" * 62)
print(" Lector de información exógena")
print("=" * 62)

if not EJEMPLO.exists():
    # El archivo NO está en el repositorio, y es a propósito.
    #
    # Es un reporte de exógena, y aunque el contribuyente de este sea
    # inventado, el repositorio es público y lo que se sube queda
    # clonado para siempre. Un reporte de exógena de una persona de
    # verdad es de lo más delicado que hay: dice dónde tiene cuentas,
    # cuánto le entró y con quién trabaja. Con uno de ejemplo adentro,
    # el día que alguien ponga ahí el de un cliente —"total, es igual
    # que el que ya está"— no habría marcha atrás.
    #
    # Así que se guarda a mano, en el computador de quien desarrolla.
    print()
    print("  No está el archivo con el que corre esta prueba:")
    print("  " + str(EJEMPLO.relative_to(RAIZ)))
    print()
    print("  Eso NO es un error del programa: ese archivo se deja fuera")
    print("  del repositorio a propósito. Un reporte de exógena dice")
    print("  dónde tiene cuentas una persona y cuánto le entró en el año,")
    print("  y el repositorio es público.")
    print()
    print("  Para correr esta prueba, ponga ahí un reporte descargado del")
    print("  portal de la DIAN con los datos cambiados por inventados, o")
    print("  pídale el suyo a quien ya lo tenga. El resto del programa")
    print("  funciona igual sin él.")
    print()
    sys.exit(0)

lectura = exogena.leer(EJEMPLO)

# ----------------------------------------------------------
print("\nA. Encuentra la tabla sola, sin filas escritas a mano")
# ----------------------------------------------------------
revisar("la hoja es la del reporte", lectura["hoja"] == "Reporte", lectura["hoja"])
revisar("encontró la fila de encabezados buscando el texto",
        lectura["fila_encabezado"] == 14, lectura["fila_encabezado"])

# Y si el archivo NO tiene esa columna, tiene que fallar con un mensaje
# que el contador entienda, no con un error de programador.
temporal = Path(tempfile.mkdtemp(prefix="taxi-exogena-"))
sin_columna = temporal / "sin_uso_sugerido.xlsx"
libro = load_workbook(EJEMPLO)
libro["Reporte"].cell(row=14, column=7).value = "Otra cosa"
libro.save(sin_columna)

try:
    exogena.leer(sin_columna)
    revisar("avisa cuando falta la columna de uso sugerido", False, "no falló")
except exogena.ExogenaInvalida as error:
    mensaje = str(error)
    revisar("avisa cuando falta la columna de uso sugerido",
            "Uso declaración Sugerida" in mensaje, mensaje[:60])
    revisar("y el mensaje está escrito para el contador",
            "DIAN" in mensaje and "Traceback" not in mensaje, mensaje[:60])

no_es_excel = temporal / "cualquier_cosa.xlsx"
no_es_excel.write_bytes(b"esto no es un excel")
try:
    exogena.leer(no_es_excel)
    revisar("un archivo que no es Excel se rechaza", False, "no falló")
except exogena.ExogenaInvalida as error:
    revisar("un archivo que no es Excel se rechaza", True, str(error)[:60])

shutil.rmtree(temporal, ignore_errors=True)

# ----------------------------------------------------------
print("\nB. La cabecera: de quién es y hasta cuándo llega")
# ----------------------------------------------------------
cabecera = lectura["cabecera"]
revisar("el año gravable", cabecera["anio"] == "2025", cabecera["anio"])
revisar("el tipo de documento", cabecera["tipo_documento"] == "C. C.",
        cabecera["tipo_documento"])
revisar("la identificación, como texto y sin decimales",
        cabecera["identificacion"].isdigit()
        and "." not in cabecera["identificacion"],
        cabecera["identificacion"])
revisar("el nombre del contribuyente",
        len(cabecera["nombre"]) > 3, cabecera["nombre"])
revisar("la fecha de corte del proceso",
        cabecera["fecha_corte"].startswith("2026-08-26"), cabecera["fecha_corte"])
revisar("y la fecha en que se generó el reporte, que es otra",
        cabecera["fecha_reporte"].startswith("2026-08-26"),
        cabecera["fecha_reporte"])

# ----------------------------------------------------------
print("\nC. Los tres avisos legales, textuales")
# ----------------------------------------------------------
avisos = lectura["avisos"]
revisar("son tres", len(avisos) == 3, len(avisos))
revisar("el primero advierte que la información puede cambiar",
        any(a.startswith("ADVERTENCIA:") for a in avisos))
revisar("el segundo manda al tercero que reporta",
        any("presenta inconsistencias" in a for a in avisos))

TERCER_AVISO = (
    "IMPORTANTE: Para cumplir con su obligación de declarar, la Información"
    " Exógena Tributaria NO ES INDISPENSABLE y NO REEMPLAZA la información de"
    " su realidad económica, ni lo exonera de declarar los valores totales que"
    " correspondan y que son de su conocimiento exclusivo."
)
revisar("y el tercero está palabra por palabra, sin resumir",
        any(a.strip() == TERCER_AVISO for a in avisos))

# ----------------------------------------------------------
print("\nD. Los cinco topes van aparte: son resumen, no renglones")
# ----------------------------------------------------------
topes = lectura["topes"]
revisar("son cinco", len(topes) == 5, len(topes))
revisar("numerados del 1 al 5",
        [t["numero"] for t in topes] == [1, 2, 3, 4, 5],
        [t["numero"] for t in topes])
esperados = {1: 36142882, 2: 129305012, 3: 42680793, 4: 168905519, 5: 11413605}
revisar("con los valores del archivo",
        {t["numero"]: t["valor"] for t in topes} == esperados)
revisar("y con el nombre que les da la DIAN",
        topes[3]["etiqueta"] == "Tope 4 - Movimiento", topes[3]["etiqueta"])
revisar("ningún tope se coló entre las filas de datos",
        not any(f["detalle"].startswith("Tope ") for f in lectura["filas"]))

# ----------------------------------------------------------
print("\nE. Las filas de datos")
# ----------------------------------------------------------
filas = lectura["filas"]
revisar("son 36 registros reportados", len(filas) == 36, len(filas))
revisar("todas traen quién reporta", all(f["nit_reporta"] for f in filas))
revisar("el NIT de quien reporta es texto, no número",
        all(isinstance(f["nit_reporta"], str) for f in filas))
revisar("sin decimales pegados ni notación científica",
        all("." not in f["nit_reporta"] and "e" not in f["nit_reporta"].lower()
            for f in filas))
# La cabecera y las filas tienen que hablar del MISMO contribuyente. Si
# no coinciden, el archivo está mezclando dos personas y eso hay que
# verlo antes de cargarlo, no después.
revisar("el NIT del contribuyente también es texto",
        all(isinstance(f["nit_contribuyente"], str) for f in filas))
revisar("y es el mismo de la cabecera en todas las filas",
        all(f["nit_contribuyente"] == cabecera["identificacion"]
            for f in filas),
        {f["nit_contribuyente"] for f in filas})
revisar("los valores son números", all(f["valor"] is not None for f in filas))
revisar("el detalle se guarda sin recortar",
        any(len(f["detalle"]) > 80 for f in filas))

por_fila = {f["fila_excel"]: f for f in filas}
revisar("una fila sin uso sugerido no rompe la lectura",
        49 in por_fila and por_fila[49]["uso_sugerido"] == "")
revisar("el uso sugerido se guarda entero, con su salto de línea",
        "\n" in por_fila[43]["uso_sugerido"])
revisar("y la nota de la DIAN queda aparte de las opciones",
        por_fila[43]["nota"].startswith("Nota:"), por_fila[43]["nota"][:40])

# ----------------------------------------------------------
print("\nF. Los códigos de concepto, que vienen dentro del detalle")
# ----------------------------------------------------------
conceptos = sorted({f["concepto"] for f in filas if f["concepto"]})
revisar("son los ocho del archivo",
        conceptos == ["1020", "1023", "1315", "1476", "2276", "4070", "5016", "5046"],
        conceptos)
revisar("«Pagos por salarios (Concepto: 2276)» da 2276",
        por_fila[32]["concepto"] == "2276", por_fila[32]["concepto"])
revisar("y una fila sin concepto lo deja vacío",
        por_fila[33]["concepto"] == "", repr(por_fila[33]["concepto"]))

# ----------------------------------------------------------
print("\nG. Los renglones del 210 que menciona la DIAN")
# ----------------------------------------------------------
todos = set()
for f in filas:
    todos |= {r["codigo"] for r in f["renglones"]}
ESPERADOS = {"R29", "R30", "R32", "R33", "R36", "R51", "R58", "R59", "R67",
             "R74", "R76", "R84", "R100", "R131", "R132"}
revisar("están los quince códigos", todos == ESPERADOS,
        sorted(todos - ESPERADOS) or sorted(ESPERADOS - todos) or "iguales")
revisar("R100 se lee como R100 y no como R10",
        "R100" in todos and "R10" not in todos)
revisar("una fila con un solo renglón trae uno",
        [r["codigo"] for r in por_fila[22]["renglones"]] == ["R132"])
revisar("una fila con varios los trae todos",
        [r["codigo"] for r in por_fila[43]["renglones"]]
        == ["R29", "R32", "R36", "R51", "R67", "R84"])
revisar("y los que van pegados con barras también",
        [r["codigo"] for r in por_fila[25]["renglones"]] == ["R33", "R59", "R100"])

revisar("los topes citados en el uso sugerido se leen por número",
        [t["numero"] for t in por_fila[26]["topes"]] == [1],
        [t["numero"] for t in por_fila[26]["topes"]])

# ----------------------------------------------------------
print("\nH. Las filas que requieren decisión del contador")
# ----------------------------------------------------------
decisiones = [f for f in filas if f["requiere_decision"]]
revisar("son ocho", len(decisiones) == 8, len(decisiones))
revisar("y son las que traen más de un renglón posible",
        all(len(f["renglones"]) > 1 for f in decisiones))
revisar("una fila con un solo renglón NO pide decisión",
        not por_fila[32]["requiere_decision"])

# Las opciones se le muestran al contador tal como la DIAN las escribió.
# Ni se reescriben, ni se les arregla el doble espacio.
opciones_24 = por_fila[24]["opciones"]
revisar("las opciones van palabra por palabra",
        opciones_24 == ["R29 Patrimonio Bruto (si el saldo es positivo)",
                        "R30 Deudas  (si el saldo es negativo)"],
        opciones_24)
revisar("con el doble espacio de la DIAN incluido",
        "Deudas  (si" in opciones_24[1])

# ----------------------------------------------------------
print("\nI. Los posibles duplicados: se marcan, no se resuelven")
# ----------------------------------------------------------
parejas = {tuple(d["filas"]): d for d in lectura["duplicados"]}

revisar("las cesantías del empleador y las del fondo se marcan",
        (43, 47) in parejas)
if (43, 47) in parejas:
    revisar("con confianza alta, porque son terceros distintos",
            parejas[(43, 47)]["confianza"] == "alta")
    revisar("y el motivo dice por qué se marcó",
            "Otro tercero" in parejas[(43, 47)]["motivo"],
            parejas[(43, 47)]["motivo"][:50])

revisar("el mismo tercero informando dos veces 114.465 se marca",
        (23, 51) in parejas)
revisar("2.460.998 y 2.460.907 se parecen en menos del 1%",
        por_fila[43]["posible_duplicado"] and por_fila[47]["posible_duplicado"])
revisar("una fila sin pareja no queda marcada",
        not por_fila[32]["posible_duplicado"])
revisar("cada marca dice con cuál fila se parece",
        por_fila[47]["duplicado_de"][0]["fila_excel"] == 43,
        por_fila[47]["duplicado_de"])
revisar("y ninguna fila se unió ni se descartó",
        len(lectura["filas"]) == 36)

# Salud y pensión valen lo mismo y las reporta el mismo empleador. Se
# marcan —el programa no sabe de impuestos— pero con confianza media y
# diciendo por qué, para que el contador lo descarte de un vistazo.
revisar("salud y pensión se marcan con confianza media, no alta",
        (25, 31) in parejas and parejas[(25, 31)]["confianza"] == "media")

# ----------------------------------------------------------
print("\nJ. El nombre de cada renglón sale del archivo, no de la IA")
# ----------------------------------------------------------
catalogo = {r["codigo"]: r for r in lectura["renglones"]}
revisar("hay un renglón por código", len(catalogo) == 15, len(catalogo))
revisar("todos tienen nombre", all(r["nombre"] for r in catalogo.values()),
        [c for c, r in catalogo.items() if not r["nombre"]])
revisar("R32 queda como lo escribe la DIAN",
        catalogo["R32"]["titulo"]
        == "R32 — Ingresos brutos por rentas de trabajo (art. 103 E.T.)",
        catalogo["R32"]["titulo"])
revisar("R29 toma el nombre corto, no el que trae la condición",
        catalogo["R29"]["nombre"] == "Patrimonio Bruto",
        catalogo["R29"]["nombre"])
revisar("R36 no se traga la frase que menciona R51",
        catalogo["R36"]["nombre"] == "Otras rentas exentas",
        catalogo["R36"]["nombre"])
revisar("R33 usa la etiqueta que la DIAN le puso entre paréntesis",
        catalogo["R33"]["nombre"] == "Trabajo", catalogo["R33"]["nombre"])
revisar("y ningún nombre lleva otro código adentro",
        not any(any(x in r["nombre"] for x in ("R2", "R3", "R5", "R6", "R7", "R8"))
                for r in catalogo.values()))

print()
print("=" * 62)
print(" %d de %d comprobaciones pasaron." % (TOTAL[0] - len(fallos), TOTAL[0]))
print(" Todo bien." if not fallos else " HAY FALLAS.")
print("=" * 62)
sys.exit(1 if fallos else 0)
